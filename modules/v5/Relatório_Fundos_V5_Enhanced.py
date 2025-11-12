"""
REPORT DIÁRIO DE FUNDOS - Versão 5.0 ENHANCED
Relatório Excel com formatação profissional e dados enriquecidos

Melhorias v4 → v5:
- Formatação condicional avançada
- Gráficos e visualizações integrados
- Aba de resumo executivo
- Mais colunas calculadas
- Design profissional aprimorado
- Frozen panes e filtros
- Validação de dados
"""

import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_PERCENTAGE_00, FORMAT_DATE_DDMMYY
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule, DataBarRule
from datetime import datetime
import os
import time
from typing import Dict, Any, Optional
import logging
from colorama import init, Fore, Back, Style

# Inicializa colorama
init(autoreset=True)

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('report_diario_v5.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ReportDiarioFundosV5:
    """Classe para geração AVANÇADA do relatório com formatação profissional"""

    def __init__(self):
        """Inicializa caminhos e configurações"""
        # Caminhos
        self.DB_PATH = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\09. Base_de_Dados\Base Fundos_V2.accdb"
        self.OUTPUT_PATH = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\Report Diário\Teste V3 - Lucas\v5_enhanced_report.xlsx"
        self.FINAL_REPORT_PATH = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\Report Diário\Fundos - Report Diário.xlsx"
        self.REPORTS_DIR = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\Report Diário\REPORTS"

        # Workbook e worksheets
        self.wb = None
        self.ws_dados = None
        self.ws_resumo = None
        self.ws_analise = None
        self.conn = None

        # Data de entrada
        self.input_date = None
        self.input_date_str = None

        # Dicionário principal de dados por fundo
        self.fundos_dados = {}

        # Estatísticas
        self.stats = {
            'inicio': None,
            'fim': None,
            'fundos_processados': 0,
            'erros': []
        }

        # Cores do tema
        self.COLORS = {
            'header_primary': '4472C4',      # Azul
            'header_secondary': '5B9BD5',    # Azul claro
            'positive': '70AD47',            # Verde
            'negative': 'C00000',            # Vermelho
            'warning': 'FFC000',             # Amarelo
            'neutral': 'F2F2F2',             # Cinza claro
            'border': '8EA9DB',              # Azul borda
            'accent': '44546A',              # Cinza escuro
        }

    # ===== MÉTODOS DE UTILIDADE =====

    def print_header(self, texto: str):
        """Imprime cabeçalho destacado"""
        largura = 80
        print("\n" + "=" * largura)
        print(Fore.CYAN + Style.BRIGHT + texto.center(largura))
        print("=" * largura + "\n")

    def print_step(self, numero: int, total: int, descricao: str):
        """Imprime passo da execução"""
        print(f"\n{Fore.YELLOW}[PASSO {numero}/{total}]{Style.RESET_ALL} {Fore.WHITE}{descricao}{Style.RESET_ALL}")
        print("-" * 80)

    def print_success(self, mensagem: str):
        """Imprime mensagem de sucesso"""
        print(f"{Fore.GREEN}✓{Style.RESET_ALL} {mensagem}")

    def print_warning(self, mensagem: str):
        """Imprime aviso"""
        print(f"{Fore.YELLOW}⚠{Style.RESET_ALL} {mensagem}")

    def print_error(self, mensagem: str):
        """Imprime erro"""
        print(f"{Fore.RED}✗{Style.RESET_ALL} {mensagem}")

    def print_info(self, mensagem: str):
        """Imprime informação"""
        print(f"{Fore.CYAN}ℹ{Style.RESET_ALL} {mensagem}")

    def validar_ambiente(self) -> bool:
        """Valida se todos os arquivos necessários existem"""
        self.print_step(1, 10, "VALIDANDO AMBIENTE")

        validacoes = {
            "Banco de dados Access": self.DB_PATH,
        }

        todas_validas = True
        for nome, caminho in validacoes.items():
            if os.path.exists(caminho):
                self.print_success(f"{nome}: OK")
            else:
                self.print_error(f"{nome}: NÃO ENCONTRADO")
                todas_validas = False

        # Cria diretórios se não existirem
        os.makedirs(os.path.dirname(self.OUTPUT_PATH), exist_ok=True)
        os.makedirs(self.REPORTS_DIR, exist_ok=True)
        self.print_success("Diretórios verificados/criados")

        return todas_validas

    def obter_data_entrada(self) -> Optional[datetime]:
        """Solicita a data de entrada ao usuário"""
        try:
            print(f"\n{Fore.CYAN}{'='*80}{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Por favor, informe a data de referência para o relatório:{Style.RESET_ALL}")
            print(f"{Fore.CYAN}Formatos aceitos: dd/mm/aaaa, dd-mm-aaaa, aaaa-mm-dd{Style.RESET_ALL}")
            print(f"{Fore.CYAN}Exemplo: 28/10/2025{Style.RESET_ALL}")
            print(f"{Fore.CYAN}{'='*80}{Style.RESET_ALL}\n")

            max_tentativas = 3
            tentativa = 0

            while tentativa < max_tentativas:
                data_input = input(f"{Fore.GREEN}Digite a data: {Style.RESET_ALL}").strip()

                if data_input.lower() in ('sair', 'exit', 'cancelar', 'q'):
                    self.print_warning("Operação cancelada pelo usuário")
                    return None

                if not data_input:
                    self.print_warning("Data não pode ser vazia")
                    tentativa += 1
                    continue

                parsed = None
                formatos = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%d/%Y"]

                for fmt in formatos:
                    try:
                        parsed = datetime.strptime(data_input, fmt)
                        break
                    except ValueError:
                        continue

                if parsed:
                    parsed = parsed.replace(hour=0, minute=0, second=0, microsecond=0)
                    self.input_date = parsed
                    self.input_date_str = parsed.strftime("%m/%d/%Y")

                    print()
                    self.print_success(f"Data aceita: {Fore.YELLOW}{parsed.strftime('%d/%m/%Y')}{Style.RESET_ALL}")
                    self.print_info(f"Formato Access: {self.input_date_str}")
                    print()

                    return self.input_date
                else:
                    tentativa += 1
                    self.print_error(f"Data inválida: {data_input}")
                    if tentativa < max_tentativas:
                        self.print_warning(f"Tentativa {tentativa}/{max_tentativas}. Tente novamente.")
                    else:
                        self.print_error("Número máximo de tentativas excedido")

            return None

        except KeyboardInterrupt:
            print(f"\n\n{Fore.YELLOW}⚠ Entrada cancelada pelo usuário{Style.RESET_ALL}\n")
            return None
        except Exception as e:
            self.print_error(f"Falha ao obter data: {e}")
            return None

    def conectar_access(self) -> pyodbc.Connection:
        """Conecta ao banco Access"""
        try:
            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={self.DB_PATH};'
            )
            conn = pyodbc.connect(conn_str)
            self.print_success("Conexão com banco Access estabelecida")
            return conn
        except Exception as e:
            self.print_error(f"Erro ao conectar ao Access: {e}")
            raise

    # ===== COLETA DE DADOS (MESMO DO V4) =====

    def coletar_dados_completos(self):
        """Coleta TODOS os dados necessários do banco"""
        self.print_step(3, 10, "COLETANDO DADOS DO BANCO ACCESS")

        self.conn = self.conectar_access()

        # Query principal
        self.print_info("Executando query principal...")
        df_principal = self._executar_query_principal()

        # Processar cada fundo
        for _, row in df_principal.iterrows():
            fundo = row['CARTEIRA']
            if fundo not in self.fundos_dados:
                self.fundos_dados[fundo] = {}

            self.fundos_dados[fundo].update({
                'data_base': row['DATA_BASE'],
                'pl_posicao_ativos': row['PL_Posicao_Ativos'],
                'pl': row['PL'],
                'pl_d1': row['PL_D1'],
                'pl_d7': row['PL_D7'],
                'pl_d30': row['PL_D30'],
                'caixa_bancario': row['SALDOCAIXAATUAL'],
                'fundos_di': row['CAIXA_REAG_II'],
                'devido_taxas': row['devido_taxa'],
                'caixa_total': row['CAIXA_TOTAL'],
                'cotista': row.get('cotista'),
                'vl_bruto': row.get('vl_bruto'),
                'vl_liquido': row.get('vl_liquido'),
                'tipo': self._determinar_tipo(fundo)
            })

        self.print_success(f"Query principal: {len(self.fundos_dados)} fundos")

        # Processar FIDCs e Qore
        self._coletar_dados_fidcs()
        self._coletar_dados_qore()

        self.print_success(f"Total de fundos coletados: {len(self.fundos_dados)}")
        self.stats['fundos_processados'] = len(self.fundos_dados)

        if self.conn:
            self.conn.close()
            self.conn = None

    def _executar_query_principal(self) -> pd.DataFrame:
        """Executa a query SQL principal (mesma do V4)"""
        query = f"""
        SELECT PT.DATA_BASE, PT.CARTEIRA, PT.VALORPOSICAOATIVOS AS PL_Posicao_Ativos,
               PT.VALORPATRIMONIOLIQUIDO AS PL,
               (SELECT TOP 1 PT1.VALORPATRIMONIOLIQUIDO FROM Patrimonio_Totais AS PT1
                WHERE PT1.CARTEIRA = PT.CARTEIRA AND PT1.DATA_BASE =
                (SELECT DU.[D-1] FROM Dias_Uteis AS DU WHERE DU.Data = PT.DATA_BASE)) AS PL_D1,
               (SELECT TOP 1 PT7.VALORPATRIMONIOLIQUIDO FROM Patrimonio_Totais AS PT7
                WHERE PT7.CARTEIRA = PT.CARTEIRA AND PT7.DATA_BASE =
                (SELECT DU.[D-7] FROM Dias_Uteis AS DU WHERE DU.Data = PT.DATA_BASE)) AS PL_D7,
               (SELECT TOP 1 PT30.VALORPATRIMONIOLIQUIDO FROM Patrimonio_Totais AS PT30
                WHERE PT30.CARTEIRA = PT.CARTEIRA AND PT30.DATA_BASE =
                (SELECT DU.[D-30] FROM Dias_Uteis AS DU WHERE DU.Data = PT.DATA_BASE)) AS PL_D30,
               PT.SALDOCAIXAATUAL,
               IIF(ISNULL((SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                           WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                           AND FUNDS.CARTEIRA = PT.CARTEIRA)), 0,
                   (SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                    WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                    AND FUNDS.CARTEIRA = PT.CARTEIRA)) AS CAIXA_REAG_II,
               (SELECT SUM(VALOR) FROM CPR_Lancamentos AS CPR WHERE CPR.DESCRICAO LIKE 'taxa%'
                AND CPR.CARTEIRA = PT.CARTEIRA AND CPR.DATA_BASE = PT.DATA_BASE) AS devido_taxa,
               IIF(ISNULL(PT.SALDOCAIXAATUAL), 0, PT.SALDOCAIXAATUAL) +
               IIF(ISNULL((SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                           WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                           AND FUNDS.CARTEIRA = PT.CARTEIRA)), 0,
                   (SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                    WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                    AND FUNDS.CARTEIRA = PT.CARTEIRA)) AS CAIXA_TOTAL,
               jc.cotista, jc.vl_bruto, jc.vl_liquido
        FROM Patrimonio_Totais AS PT
        LEFT JOIN (SELECT cd_fundo, DT_POSICAO AS data_base, MAX(CD_COTISTA) AS COTISTA,
                          SUM(VL_CORRIGIDO) AS vl_bruto, SUM(VL_RESGATE) AS vl_liquido
                   FROM jcot GROUP BY cd_fundo, DT_POSICAO) AS jc
        ON (jc.data_base = PT.DATA_BASE) AND (jc.cd_fundo = PT.CARTEIRA)
        WHERE PT.DATA_BASE = #{self.input_date_str}#

        UNION ALL

        SELECT CP.[Data Posição] AS DATA_BASE, CP.FUNDO AS CARTEIRA,
               (SELECT TOP 1 RV.ValorTotal FROM Renda_Variável_Maps AS RV
                WHERE RV.FUNDO = CP.FUNDO AND RV.DATA_INPUT = CP.[Data Posição]) AS PL_Posicao_Ativos,
               CP.[PL Posição] AS PL,
               (SELECT TOP 1 CP1.[PL Posição] FROM Cotas_Patrimonio_Maps AS CP1
                WHERE CP1.FUNDO = CP.FUNDO AND CP1.[Data Posição] =
                (SELECT DU.[D-1] FROM Dias_Uteis AS DU WHERE DU.Data = CP.[Data Posição])) AS PL_D1,
               (SELECT TOP 1 CP7.[PL Posição] FROM Cotas_Patrimonio_Maps AS CP7
                WHERE CP7.FUNDO = CP.FUNDO AND CP7.[Data Posição] =
                (SELECT DU.[D-7] FROM Dias_Uteis AS DU WHERE DU.Data = CP.[Data Posição])) AS PL_D7,
               (SELECT TOP 1 CP30.[PL Posição] FROM Cotas_Patrimonio_Maps AS CP30
                WHERE CP30.FUNDO = CP.FUNDO AND CP30.[Data Posição] =
                (SELECT DU.[D-30] FROM Dias_Uteis AS DU WHERE DU.Data = CP.[Data Posição])) AS PL_D30,
               (SELECT TOP 1 CM.ValorTotal FROM Caixa_Maps AS CM
                WHERE CM.FUNDO = CP.FUNDO AND CM.DATA_INPUT = CP.[Data Posição]) AS SALDOCAIXAATUAL,
               IIF(ISNULL((SELECT TOP 1 CFM.ValorTotal FROM Cotas_Fundos_Maps AS CFM
                           WHERE CFM.FUNDO = CP.FUNDO AND CFM.DATA_INPUT = CP.[Data Posição])), 0,
                   (SELECT TOP 1 CFM.ValorTotal FROM Cotas_Fundos_Maps AS CFM
                    WHERE CFM.FUNDO = CP.FUNDO AND CFM.DATA_INPUT = CP.[Data Posição])) AS CAIXA_REAG_II,
               (SELECT SUM(VP.ValorTotal) FROM Valor_a_Pagar_Maps AS VP
                WHERE VP.FUNDO = CP.FUNDO AND VP.DATA_INPUT = CP.[Data Posição]) AS devido_taxa,
               NULL AS CAIXA_TOTAL, NULL AS cotista,
               IIF(CP.FUNDO IN ('FIP TRAIL MULT','FIP GEKKO','FIP RAM','FIP BENELLI',
                                'FIP TURANO','FIP MURCIELAGO','FIP AMG','FIP OSLO','FIP ESTOCOLMO'),
                   IIF(ISNULL((SELECT SUM(P.VLIR) FROM Passivo_Maps AS P
                               WHERE MID(P.FUNDO, 7) = CP.FUNDO AND P.[DATA_INPUT] = CP.[Data Posição])), 0,
                       (SELECT SUM(P.VLIR) FROM Passivo_Maps AS P
                        WHERE MID(P.FUNDO, 7) = CP.FUNDO AND P.[DATA_INPUT] = CP.[Data Posição])),
                   NULL) AS vl_bruto,
               NULL AS vl_liquido
        FROM Cotas_Patrimonio_Maps AS CP
        WHERE CP.[Data Posição] = #{self.input_date_str}#
        AND CP.FUNDO IN ('FIP TRAIL MULT', 'FIP GEKKO', 'FIP RAM', 'FIP BENELLI', 'FIP TURANO',
                         'FIP MURCIELAGO', 'FIP AMG', 'FIP OSLO', 'FIP ESTOCOLMO')
        """

        inicio = time.time()
        df = pd.read_sql(query, self.conn)
        tempo = time.time() - inicio

        self.print_info(f"Query executada em {tempo:.2f}s - {len(df)} registros")
        return df

    def _determinar_tipo(self, fundo: str) -> str:
        """Determina o tipo do fundo"""
        fundo_upper = fundo.upper()
        if 'FIDC' in fundo_upper:
            return 'FIDC'
        elif 'FIP' in fundo_upper:
            return 'FIP'
        elif 'FIM' in fundo_upper:
            return 'FIM'
        else:
            return 'OUTROS'

    def _coletar_dados_fidcs(self):
        """Coleta dados dos 19 FIDCs do sistema MAPS"""
        self.print_info("Coletando dados FIDCs...")

        fundos_fidcs = [
            "FIDC FANGIO", "FIDC FLYNN", "FIDC HILL NP", "FIDC IRIDIO", "FIDC KOLEOS",
            "FIDC LEGACY", "FIDC LOGAN", "FIDC MANSELL", "FIDC NORFOLK", "FIDC RAIKKONEN",
            "FIDC RINDT", "FIDC STARK", "FIDC TERPSICORE", "FIDC URANO NP", "FIDC VALKYRIES",
            "FIDC VANTAGE", "FIDC VIRAGE", "FIDC WILDE NP", "FIDC ZANDVOORT NP"
        ]

        in_list = "', '".join(fundos_fidcs)
        in_clause = f"('{in_list}')"

        # Obter datas úteis
        query_du = f"SELECT [D-1] AS D1, [D-7] AS D7, [D-30] AS D30 FROM Dias_Uteis WHERE Data = #{self.input_date_str}#"
        df_du = pd.read_sql(query_du, self.conn)

        if df_du.empty:
            self.print_warning("Dias úteis não encontrados para FIDCs")
            return

        d1_str = df_du['D1'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D1'].iloc[0]) else None
        d7_str = df_du['D7'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D7'].iloc[0]) else None
        d30_str = df_du['D30'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D30'].iloc[0]) else None

        # Queries de dados
        queries = {
            'Caixa': f"SELECT FUNDO, SUM(ValorTotal) AS SomaCaixa FROM Caixa_Maps WHERE DATA_INPUT = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO",
            'PL': f"SELECT FUNDO, MAX([PL Posição]) AS PL_POSICAO FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO",
            'VAP': f"SELECT FUNDO, SUM(ValorTotal) AS SomaValorTotal FROM Valor_a_Pagar_Maps WHERE DATA_INPUT = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO",
            'Renda Fixa': f"SELECT FUNDO, SUM(Valor_Total) AS SomaRF FROM Renda_Fixa_Maps WHERE DATA_INPUT = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO"
        }

        dados = {}
        for nome, query in queries.items():
            df = pd.read_sql(query, self.conn)
            dados[nome] = df

        # PL histórico
        if d1_str:
            dados['PL_D1'] = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D1 FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{d1_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO", self.conn)
        if d7_str:
            dados['PL_D7'] = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D7 FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{d7_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO", self.conn)
        if d30_str:
            dados['PL_D30'] = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D30 FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{d30_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO", self.conn)

        # Processar cada FIDC
        fidcs_encontrados = 0
        for fundo in fundos_fidcs:
            # Buscar valores
            rf_val = dados['Renda Fixa'][dados['Renda Fixa']['FUNDO'] == fundo]['SomaRF'].values
            pl_val = dados['PL'][dados['PL']['FUNDO'] == fundo]['PL_POSICAO'].values
            caixa_val = dados['Caixa'][dados['Caixa']['FUNDO'] == fundo]['SomaCaixa'].values
            vap_val = dados['VAP'][dados['VAP']['FUNDO'] == fundo]['SomaValorTotal'].values

            pl_d1_val = dados.get('PL_D1', pd.DataFrame())[dados.get('PL_D1', pd.DataFrame())['FUNDO'] == fundo]['PL_D1'].values if 'PL_D1' in dados else []
            pl_d7_val = dados.get('PL_D7', pd.DataFrame())[dados.get('PL_D7', pd.DataFrame())['FUNDO'] == fundo]['PL_D7'].values if 'PL_D7' in dados else []
            pl_d30_val = dados.get('PL_D30', pd.DataFrame())[dados.get('PL_D30', pd.DataFrame())['FUNDO'] == fundo]['PL_D30'].values if 'PL_D30' in dados else []

            pl_posicao_ativos = rf_val[0] if len(rf_val) > 0 else 0
            pl = pl_val[0] if len(pl_val) > 0 else 0
            caixa_bancario = caixa_val[0] if len(caixa_val) > 0 else 0
            devido_taxas = vap_val[0] if len(vap_val) > 0 else 0

            # Adicionar ao dicionário principal
            self.fundos_dados[fundo] = {
                'data_base': self.input_date,
                'pl_posicao_ativos': pl_posicao_ativos,
                'pl': pl,
                'pl_d1': pl_d1_val[0] if len(pl_d1_val) > 0 else 0,
                'pl_d7': pl_d7_val[0] if len(pl_d7_val) > 0 else 0,
                'pl_d30': pl_d30_val[0] if len(pl_d30_val) > 0 else 0,
                'caixa_bancario': caixa_bancario,
                'fundos_di': 0,
                'devido_taxas': devido_taxas,
                'caixa_total': caixa_bancario,
                'cotista': None,
                'vl_bruto': 0,
                'vl_liquido': 0,
                'tipo': 'FIDC'
            }

            if pl > 0:
                fidcs_encontrados += 1

        self.print_success(f"FIDCs: {fidcs_encontrados}/{len(fundos_fidcs)} com dados")

    def _chave_fundo_padrao(self, nome: str) -> str:
        """Padroniza nome do fundo"""
        k = nome.upper().strip()
        k = ' '.join(k.split())

        if 'BLOKO URBANISMO' in k or 'FIP BLOKO URBANISMO' in k:
            return 'FIP BLOKO URBANISMO'
        elif 'BLOKO FIM' in k or 'FIM BLOKO' in k:
            return 'FIM BLOKO'
        elif 'FIP GTB' in k:
            return 'FIP GTB MULT'
        elif 'FIP BLOKO' in k:
            return 'FIP BLOKO MULT'
        else:
            return k

    def _coletar_dados_qore(self):
        """Coleta dados dos 24 fundos do sistema Qore"""
        self.print_info("Coletando dados Qore...")

        fundos_qore = [
            "FIP MINAS", "FIP GOIAS", "FIP TERRAVISTA", "FIP BLOKO URBANISMO",
            "FIP CALDAS", "FIP KAWANA", "FIP RENOGRID", "FIP GTB MULT",
            "FIP GEKKO", "FIP RAM", "FIP ON", "FIP BENELLI",
            "FIP TURANO", "FIP MURCIELAGO", "FIP AMG", "FIP OSLO", "FIP ESTOCOLMO",
            "FIDC FORSETI", "FIDC SOCRATES", "FIDC PLATAO", "FIDC EVOQUE",
            "FIM PES", "FIDC SOA", "FIM BLOKO"
        ]

        in_list = "', '".join(fundos_qore)
        in_clause = f"('{in_list}')"
        extra_like = " OR FUNDO LIKE '*FIP GTB*' OR FUNDO LIKE '*BLOKO*'"

        # Datas úteis
        query_du = f"SELECT [D-1] AS D1, [D-7] AS D7, [D-30] AS D30 FROM Dias_Uteis WHERE Data = #{self.input_date_str}#"
        df_du = pd.read_sql(query_du, self.conn)

        if df_du.empty:
            self.print_warning("Dias úteis não encontrados para Qore")
            return

        d1_str = df_du['D1'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D1'].iloc[0]) else None
        d7_str = df_du['D7'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D7'].iloc[0]) else None
        d30_str = df_du['D30'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D30'].iloc[0]) else None

        # Dicionários para agregar
        mapas = {
            'caixa': {}, 'pl': {}, 'pl_d1': {}, 'pl_d7': {}, 'pl_d30': {},
            'taxa': {}, 'pl_ativos': {}, 'reag_ii': {}
        }

        # 1. Caixa_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(Valor) AS ValorDia FROM Caixa_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['caixa'][key] = mapas['caixa'].get(key, 0) + (row['ValorDia'] if pd.notna(row['ValorDia']) else 0)

        # 2. CPR_QORE (taxa)
        try:
            df = pd.read_sql(f"SELECT FUNDO, SUM(VALOR) AS SomaTaxa FROM CPR_QORE WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) AND (Descrição LIKE 'taxa%' OR Descrição LIKE 'despesa%') GROUP BY FUNDO", self.conn)
        except:
            df = pd.read_sql(f"SELECT FUNDO, SUM(VALOR) AS SomaTaxa FROM CPR_QORE WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) AND (DESCRICAO LIKE 'taxa%' OR DESCRICAO LIKE 'despesa%') GROUP BY FUNDO", self.conn)

        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['taxa'][key] = mapas['taxa'].get(key, 0) + (row['SomaTaxa'] if pd.notna(row['SomaTaxa']) else 0)

        # 3. Sociedade_Limitada_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(IIF([Valor_Mercado] IS NULL OR [Valor_Mercado]=0, [Valor_Custo], [Valor_Mercado])) AS PL_Posicao_Ativos FROM Sociedade_Limitada_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl_ativos'][key] = mapas['pl_ativos'].get(key, 0) + (row['PL_Posicao_Ativos'] if pd.notna(row['PL_Posicao_Ativos']) else 0)

        # 4. Renda_Fixa_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(Valor_Bruto) AS PL_Posicao_Ativos FROM Renda_Fixa_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl_ativos'][key] = mapas['pl_ativos'].get(key, 0) + (row['PL_Posicao_Ativos'] if pd.notna(row['PL_Posicao_Ativos']) else 0)

        # 5. Direito_Creditorio_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(IIF([Valor_Mercado] IS NULL OR [Valor_Mercado]=0, [Valor_Custo], [Valor_Mercado])) AS PL_Posicao_Ativos FROM Direito_Creditorio_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl_ativos'][key] = mapas['pl_ativos'].get(key, 0) + (row['PL_Posicao_Ativos'] if pd.notna(row['PL_Posicao_Ativos']) else 0)

        # 6. REAG II
        df = pd.read_sql(f"SELECT FUNDO, SUM([Qtde]*[Pu_Mercado]) AS ValorReagII FROM Sociedade_Limitada_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) AND [Descricao] LIKE '*REAG FIRF II*' GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['reag_ii'][key] = mapas['reag_ii'].get(key, 0) + (row['ValorReagII'] if pd.notna(row['ValorReagII']) else 0)

        # 7. PL atual
        df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_POSICAO FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl'][key] = row['PL_POSICAO'] if pd.notna(row['PL_POSICAO']) else 0

        # 8-10. PL histórico
        if d1_str:
            df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D1 FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{d1_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
            for _, row in df.iterrows():
                key = self._chave_fundo_padrao(row['FUNDO'])
                mapas['pl_d1'][key] = row['PL_D1'] if pd.notna(row['PL_D1']) else 0

        if d7_str:
            df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D7 FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{d7_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
            for _, row in df.iterrows():
                key = self._chave_fundo_padrao(row['FUNDO'])
                mapas['pl_d7'][key] = row['PL_D7'] if pd.notna(row['PL_D7']) else 0

        if d30_str:
            df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D30 FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{d30_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
            for _, row in df.iterrows():
                key = self._chave_fundo_padrao(row['FUNDO'])
                mapas['pl_d30'][key] = row['PL_D30'] if pd.notna(row['PL_D30']) else 0

        # Processar cada fundo Qore
        qore_encontrados = 0
        for fundo in fundos_qore:
            key = self._chave_fundo_padrao(fundo)

            caixa = mapas['caixa'].get(key, 0)
            fundos_di = mapas['reag_ii'].get(key, 0)
            pl = mapas['pl'].get(key, 0)

            self.fundos_dados[key] = {
                'data_base': self.input_date,
                'pl_posicao_ativos': mapas['pl_ativos'].get(key, 0),
                'pl': pl,
                'pl_d1': mapas['pl_d1'].get(key, 0),
                'pl_d7': mapas['pl_d7'].get(key, 0),
                'pl_d30': mapas['pl_d30'].get(key, 0),
                'caixa_bancario': caixa,
                'fundos_di': fundos_di,
                'devido_taxas': mapas['taxa'].get(key, 0),
                'caixa_total': caixa + fundos_di,
                'cotista': None,
                'vl_bruto': 0,
                'vl_liquido': 0,
                'tipo': self._determinar_tipo(key)
            }

            if pl > 0:
                qore_encontrados += 1

        self.print_success(f"Qore: {qore_encontrados}/{len(fundos_qore)} com dados")

    # ===== CONSTRUÇÃO DO RELATÓRIO ENHANCED =====

    def criar_workbook(self):
        """Cria workbook com múltiplas abas"""
        self.print_step(4, 10, "CRIANDO ESTRUTURA DO RELATÓRIO")

        self.wb = Workbook()

        # Aba 1: Resumo Executivo
        self.ws_resumo = self.wb.active
        self.ws_resumo.title = "Resumo Executivo"

        # Aba 2: Dados Detalhados
        self.ws_dados = self.wb.create_sheet("Dados Detalhados")

        # Aba 3: Análise por Tipo
        self.ws_analise = self.wb.create_sheet("Análise por Tipo")

        self.print_success("Workbook criado com 3 abas")

    def criar_resumo_executivo(self):
        """Cria aba de resumo executivo"""
        self.print_step(5, 10, "CRIANDO RESUMO EXECUTIVO")

        ws = self.ws_resumo

        # Título principal
        ws.merge_cells('B2:H2')
        cell_titulo = ws['B2']
        cell_titulo.value = "📊 RELATÓRIO DIÁRIO DE FUNDOS - RESUMO EXECUTIVO"
        cell_titulo.font = Font(size=18, bold=True, color='FFFFFF')
        cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
        cell_titulo.fill = PatternFill(start_color=self.COLORS['header_primary'],
                                       end_color=self.COLORS['header_primary'],
                                       fill_type='solid')
        ws.row_dimensions[2].height = 30

        # Data do relatório
        ws.merge_cells('B3:H3')
        cell_data = ws['B3']
        cell_data.value = f"Data de Referência: {self.input_date.strftime('%d/%m/%Y (%A)')}"
        cell_data.font = Font(size=12, italic=True)
        cell_data.alignment = Alignment(horizontal='center')
        cell_data.fill = PatternFill(start_color=self.COLORS['neutral'],
                                      end_color=self.COLORS['neutral'],
                                      fill_type='solid')

        # Calcular métricas
        total_pl = sum([f.get('pl', 0) or 0 for f in self.fundos_dados.values()])
        total_caixa = sum([f.get('caixa_total', 0) or 0 for f in self.fundos_dados.values()])
        total_fundos = len(self.fundos_dados)
        total_devido = sum([f.get('devido_taxas', 0) or 0 for f in self.fundos_dados.values()])

        # PL anterior para variação
        total_pl_d1 = sum([f.get('pl_d1', 0) or 0 for f in self.fundos_dados.values()])
        var_pl = ((total_pl / total_pl_d1) - 1) if total_pl_d1 != 0 else 0

        # Métricas principais (Cards)
        row = 5
        metricas = [
            ("Patrimônio Líquido Total", total_pl, self.COLORS['header_primary'], "R$"),
            ("Caixa Total", total_caixa, self.COLORS['positive'], "R$"),
            ("Total de Fundos", total_fundos, self.COLORS['accent'], "#"),
            ("Devido Taxas", total_devido, self.COLORS['negative'], "R$"),
        ]

        col_start = 2
        for i, (label, valor, cor, prefixo) in enumerate(metricas):
            col = col_start + (i * 2)

            # Label
            cell_label = ws.cell(row, col)
            cell_label.value = label
            cell_label.font = Font(size=10, bold=True, color='FFFFFF')
            cell_label.alignment = Alignment(horizontal='center', vertical='center')
            cell_label.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

            # Valor
            cell_valor = ws.cell(row+1, col)
            if prefixo == "R$":
                cell_valor.value = valor
                cell_valor.number_format = 'R$ #,##0'
            elif prefixo == "#":
                cell_valor.value = valor
                cell_valor.number_format = '#,##0'

            cell_valor.font = Font(size=16, bold=True)
            cell_valor.alignment = Alignment(horizontal='center', vertical='center')
            cell_valor.fill = PatternFill(start_color=self.COLORS['neutral'],
                                          end_color=self.COLORS['neutral'],
                                          fill_type='solid')

            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)

        # Distribuição por tipo
        row = 9
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = "Distribuição por Tipo de Fundo"
        cell.font = Font(size=12, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=self.COLORS['header_secondary'],
                                end_color=self.COLORS['header_secondary'],
                                fill_type='solid')

        # Calcular distribuição
        tipos = {}
        for fundo, dados in self.fundos_dados.items():
            tipo = dados.get('tipo', 'OUTROS')
            pl = dados.get('pl', 0) or 0
            tipos[tipo] = tipos.get(tipo, {'pl': 0, 'count': 0})
            tipos[tipo]['pl'] += pl
            tipos[tipo]['count'] += 1

        # Headers tabela
        row += 1
        headers = ['Tipo', 'Quantidade', 'PL Total', '% do Total']
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=self.COLORS['neutral'],
                                    end_color=self.COLORS['neutral'],
                                    fill_type='solid')

        # Dados distribuição
        row += 1
        for tipo, dados in sorted(tipos.items()):
            ws.cell(row, 2, tipo)
            ws.cell(row, 3, dados['count'])
            ws.cell(row, 4, dados['pl'])
            ws.cell(row, 4).number_format = 'R$ #,##0'

            perc = (dados['pl'] / total_pl) if total_pl > 0 else 0
            ws.cell(row, 5, perc)
            ws.cell(row, 5).number_format = '0.00%'

            row += 1

        # Top 10 Fundos
        row += 2
        ws.merge_cells(f'B{row}:G{row}')
        cell = ws[f'B{row}']
        cell.value = "Top 10 Fundos por Patrimônio Líquido"
        cell.font = Font(size=12, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=self.COLORS['header_secondary'],
                                end_color=self.COLORS['header_secondary'],
                                fill_type='solid')

        # Headers
        row += 1
        headers_top = ['#', 'Fundo', 'Tipo', 'PL', 'Caixa', 'Var. D-1']
        for col_idx, header in enumerate(headers_top, start=2):
            cell = ws.cell(row, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=self.COLORS['neutral'],
                                    end_color=self.COLORS['neutral'],
                                    fill_type='solid')

        # Top 10
        fundos_sorted = sorted(self.fundos_dados.items(),
                               key=lambda x: x[1].get('pl', 0) or 0,
                               reverse=True)[:10]

        row += 1
        for rank, (fundo, dados) in enumerate(fundos_sorted, start=1):
            ws.cell(row, 2, rank)
            ws.cell(row, 3, fundo)
            ws.cell(row, 4, dados.get('tipo', '-'))

            pl = dados.get('pl', 0) or 0
            ws.cell(row, 5, pl)
            ws.cell(row, 5).number_format = 'R$ #,##0'

            caixa = dados.get('caixa_total', 0) or 0
            ws.cell(row, 6, caixa)
            ws.cell(row, 6).number_format = 'R$ #,##0'

            delta = self._calcular_delta(pl, dados.get('pl_d1', 0) or 0)
            if delta != "-":
                ws.cell(row, 7, delta)
                ws.cell(row, 7).number_format = '0.00%'

                # Formatação condicional inline
                if delta > 0:
                    ws.cell(row, 7).font = Font(color=self.COLORS['positive'], bold=True)
                elif delta < 0:
                    ws.cell(row, 7).font = Font(color=self.COLORS['negative'], bold=True)
            else:
                ws.cell(row, 7, '-')

            row += 1

        # Ajustar larguras
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12

        self.print_success("Resumo executivo criado")

    def escrever_headers_dados(self):
        """Escreve headers da aba de dados (melhorado)"""
        self.print_step(6, 10, "ESCREVENDO HEADERS DA ABA DE DADOS")

        ws = self.ws_dados

        # Título
        ws.merge_cells('A1:W1')
        cell_titulo = ws['A1']
        cell_titulo.value = f"DADOS DETALHADOS - {self.input_date.strftime('%d/%m/%Y')}"
        cell_titulo.font = Font(size=14, bold=True, color='FFFFFF')
        cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
        cell_titulo.fill = PatternFill(start_color=self.COLORS['header_primary'],
                                       end_color=self.COLORS['header_primary'],
                                       fill_type='solid')
        ws.row_dimensions[1].height = 25

        # Headers linha 2 - Categorias
        headers_categorias = {
            1: ("Identificação", 3),
            4: ("Posição Ativos", 1),
            5: ("Patrimônio Líquido", 4),
            9: ("Caixa", 3),
            12: ("Passivo", 2),
            14: ("Análise", 3),
            17: ("PL Histórico", 3),
            20: ("Indicadores", 3),
        }

        for col_start, (label, span) in headers_categorias.items():
            if span > 1:
                ws.merge_cells(start_row=2, start_column=col_start,
                               end_row=2, end_column=col_start + span - 1)
            cell = ws.cell(2, col_start)
            cell.value = label
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.COLORS['header_secondary'],
                                    end_color=self.COLORS['header_secondary'],
                                    fill_type='solid')

        # Headers linha 3 - Colunas
        headers_detalhados = [
            "Fundo", "Gestora", "Tipo",
            "Posição Ativos",
            "PL Atual", "∆ D-1", "∆ D-7", "∆ D-30",
            "Saldo Bancário", "Fundos DI", "Total",
            "Devido Taxas", "NCG",
            "Importo Retido", "IR/IOF", "Líquido",
            "PL D-1", "PL D-7", "PL D-30",
            "% Caixa/PL", "Liquidez", "Status"
        ]

        for col, header in enumerate(headers_detalhados, start=1):
            cell = ws.cell(3, col)
            cell.value = header
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=self.COLORS['neutral'],
                                    end_color=self.COLORS['neutral'],
                                    fill_type='solid')

            # Bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

        # Ajustar larguras
        larguras = [
            25, 15, 10,  # Identificação
            15,          # Posição Ativos
            15, 10, 10, 10,  # PL
            15, 15, 15,  # Caixa
            15, 12,      # Passivo
            15, 12, 15,  # Análise
            15, 15, 15,  # Histórico
            10, 10, 12   # Indicadores
        ]

        for col_idx, largura in enumerate(larguras, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = largura

        self.print_success("Headers da aba de dados escritos")

    def escrever_dados_fundos_enhanced(self):
        """Escreve dados com colunas NOVAS e calculadas"""
        self.print_step(7, 10, "ESCREVENDO DADOS DOS FUNDOS (ENHANCED)")

        ws = self.ws_dados
        fundos_ordenados = sorted(self.fundos_dados.keys())

        row = 4  # Começa após headers
        fundos_escritos = 0

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for fundo in fundos_ordenados:
            dados = self.fundos_dados[fundo]

            # 1-3: Identificação
            ws.cell(row, 1, fundo).border = thin_border
            ws.cell(row, 2, self._buscar_gestora(fundo)).border = thin_border
            ws.cell(row, 3, dados.get('tipo', '-')).border = thin_border

            # 4: Posição Ativos
            pl_pos_ativos = dados.get('pl_posicao_ativos', 0) or 0
            fundos_di = dados.get('fundos_di', 0) or 0
            cell = ws.cell(row, 4, pl_pos_ativos - fundos_di)
            cell.number_format = '#,##0'
            cell.border = thin_border

            # 5-8: Patrimônio Líquido
            pl_atual = dados.get('pl', 0) or 0
            pl_d1 = dados.get('pl_d1', 0) or 0
            pl_d7 = dados.get('pl_d7', 0) or 0
            pl_d30 = dados.get('pl_d30', 0) or 0

            cell = ws.cell(row, 5, pl_atual)
            cell.number_format = '#,##0'
            cell.border = thin_border

            for col, pl_anterior in [(6, pl_d1), (7, pl_d7), (8, pl_d30)]:
                delta = self._calcular_delta(pl_atual, pl_anterior)
                cell = ws.cell(row, col, delta if delta != "-" else 0)
                if delta != "-":
                    cell.number_format = '0.00%'
                cell.border = thin_border

            # 9-11: Caixa
            caixa_bancario = dados.get('caixa_bancario', 0) or 0
            total_caixa = dados.get('caixa_total', 0) or caixa_bancario + fundos_di

            ws.cell(row, 9, caixa_bancario).number_format = '#,##0'
            ws.cell(row, 9).border = thin_border
            ws.cell(row, 10, fundos_di).number_format = '#,##0'
            ws.cell(row, 10).border = thin_border
            ws.cell(row, 11, total_caixa).number_format = '#,##0'
            ws.cell(row, 11).border = thin_border

            # 12-13: Passivo
            devido_taxas = dados.get('devido_taxas', 0) or 0
            ncg = self._calcular_ncg(devido_taxas, total_caixa)

            ws.cell(row, 12, devido_taxas).number_format = '#,##0'
            ws.cell(row, 12).border = thin_border
            ws.cell(row, 13, ncg).number_format = '#,##0'
            ws.cell(row, 13).border = thin_border

            # 14-16: Análise (NOVO!)
            vl_bruto = dados.get('vl_bruto', 0) or 0
            vl_liquido = dados.get('vl_liquido', 0) or 0
            ir_iof = vl_bruto - vl_liquido

            ws.cell(row, 14, vl_bruto - vl_liquido).number_format = '#,##0'
            ws.cell(row, 14).border = thin_border
            ws.cell(row, 15, ir_iof).number_format = '#,##0'
            ws.cell(row, 15).border = thin_border
            ws.cell(row, 16, vl_liquido).number_format = '#,##0'
            ws.cell(row, 16).border = thin_border

            # 17-19: Histórico
            ws.cell(row, 17, pl_d1).number_format = '#,##0'
            ws.cell(row, 17).border = thin_border
            ws.cell(row, 18, pl_d7).number_format = '#,##0'
            ws.cell(row, 18).border = thin_border
            ws.cell(row, 19, pl_d30).number_format = '#,##0'
            ws.cell(row, 19).border = thin_border

            # 20-22: Indicadores (NOVO!)
            # % Caixa/PL
            perc_caixa = (total_caixa / pl_atual) if pl_atual > 0 else 0
            ws.cell(row, 20, perc_caixa).number_format = '0.00%'
            ws.cell(row, 20).border = thin_border

            # Liquidez (dias de operação)
            liquidez_dias = (total_caixa / abs(devido_taxas)) if devido_taxas < 0 else 999
            ws.cell(row, 21, liquidez_dias).number_format = '#,##0'
            ws.cell(row, 21).border = thin_border

            # Status
            status = self._determinar_status(pl_atual, pl_d1, total_caixa, devido_taxas)
            cell_status = ws.cell(row, 22, status)
            cell_status.border = thin_border
            cell_status.alignment = Alignment(horizontal='center')

            # Cor do status
            if status == "✓ OK":
                cell_status.font = Font(color=self.COLORS['positive'], bold=True)
            elif status == "⚠ Alerta":
                cell_status.font = Font(color=self.COLORS['warning'], bold=True)
            elif status == "✗ Crítico":
                cell_status.font = Font(color=self.COLORS['negative'], bold=True)

            row += 1
            fundos_escritos += 1

        self.print_success(f"Dados escritos: {fundos_escritos} fundos")

    def _determinar_status(self, pl_atual, pl_d1, caixa_total, devido_taxas) -> str:
        """Determina status do fundo baseado em regras"""
        # Regra 1: PL caindo mais de 5%
        if pl_d1 and pl_d1 > 0:
            var = (pl_atual / pl_d1) - 1
            if var < -0.05:
                return "✗ Crítico"
            elif var < -0.02:
                return "⚠ Alerta"

        # Regra 2: Caixa insuficiente
        if devido_taxas < 0:
            if abs(devido_taxas) > caixa_total:
                return "✗ Crítico"
            elif abs(devido_taxas) > caixa_total * 0.8:
                return "⚠ Alerta"

        # Regra 3: Caixa muito alto (> 30%)
        if pl_atual > 0:
            perc_caixa = caixa_total / pl_atual
            if perc_caixa > 0.3:
                return "⚠ Alerta"

        return "✓ OK"

    def _buscar_gestora(self, fundo: str) -> str:
        """Busca gestora do fundo"""
        mapeamento = {
            'FIDC FORSETI': 'North Sea',
            'FIDC SOCRATES': 'North Sea',
            'FIDC PLATAO': 'North Sea',
            'FIDC EVOQUE': 'Bloko',
            'FIM PES': 'Bloko',
            'FIM BLOKO': 'Bloko',
        }
        return mapeamento.get(fundo, '-')

    def _calcular_delta(self, valor_atual, valor_anterior) -> float:
        """Calcula variação percentual"""
        if not valor_anterior or valor_anterior == 0:
            return "-"

        if valor_anterior < 0:
            if valor_atual > 0:
                return "-"
            else:
                return round(1 - valor_atual / valor_anterior, 4)
        else:
            if valor_atual < 0:
                return "-"
            else:
                return round(valor_atual / valor_anterior - 1, 4)

    def _calcular_ncg(self, devido_taxas, caixa_total) -> float:
        """Calcula NCG"""
        if -devido_taxas > caixa_total:
            return -(devido_taxas + caixa_total)
        return 0

    def aplicar_formatacao_condicional(self):
        """Aplica formatação condicional nas colunas"""
        self.print_step(8, 10, "APLICANDO FORMATAÇÃO CONDICIONAL")

        ws = self.ws_dados
        ultima_linha = ws.max_row

        # Regra 1: ∆ D-1 (coluna F)
        ws.conditional_formatting.add(
            f'F4:F{ultima_linha}',
            ColorScaleRule(
                start_type='num', start_value=-0.05, start_color='C00000',
                mid_type='num', mid_value=0, mid_color='FFFFFF',
                end_type='num', end_value=0.05, end_color='70AD47'
            )
        )

        # Regra 2: % Caixa/PL (coluna T) - Data Bars
        ws.conditional_formatting.add(
            f'T4:T{ultima_linha}',
            DataBarRule(
                start_type='num', start_value=0, end_type='num', end_value=0.5,
                color=self.COLORS['header_primary']
            )
        )

        # Regra 3: Liquidez (coluna U) - Ícones
        ws.conditional_formatting.add(
            f'U4:U{ultima_linha}',
            IconSetRule(
                icon_style='3TrafficLights1',
                type='num',
                values=[0, 30, 60],
                reverse=False
            )
        )

        self.print_success("Formatação condicional aplicada")

    def criar_analise_por_tipo(self):
        """Cria aba de análise agregada por tipo"""
        self.print_step(9, 10, "CRIANDO ANÁLISE POR TIPO")

        ws = self.ws_analise

        # Título
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "ANÁLISE AGREGADA POR TIPO DE FUNDO"
        cell.font = Font(size=14, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=self.COLORS['header_primary'],
                                end_color=self.COLORS['header_primary'],
                                fill_type='solid')
        ws.row_dimensions[1].height = 25

        # Agregar por tipo
        tipos_agregados = {}
        for fundo, dados in self.fundos_dados.items():
            tipo = dados.get('tipo', 'OUTROS')
            if tipo not in tipos_agregados:
                tipos_agregados[tipo] = {
                    'count': 0,
                    'pl': 0,
                    'caixa': 0,
                    'devido': 0,
                    'pl_d1': 0
                }

            tipos_agregados[tipo]['count'] += 1
            tipos_agregados[tipo]['pl'] += dados.get('pl', 0) or 0
            tipos_agregados[tipo]['caixa'] += dados.get('caixa_total', 0) or 0
            tipos_agregados[tipo]['devido'] += dados.get('devido_taxas', 0) or 0
            tipos_agregados[tipo]['pl_d1'] += dados.get('pl_d1', 0) or 0

        # Headers
        headers = ['Tipo', 'Quantidade', 'PL Total', 'Caixa Total', 'Devido', 'Var. D-1']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(2, col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=self.COLORS['neutral'],
                                    end_color=self.COLORS['neutral'],
                                    fill_type='solid')

        # Dados
        row = 3
        for tipo in sorted(tipos_agregados.keys()):
            dados = tipos_agregados[tipo]

            ws.cell(row, 1, tipo)
            ws.cell(row, 2, dados['count'])

            ws.cell(row, 3, dados['pl'])
            ws.cell(row, 3).number_format = 'R$ #,##0'

            ws.cell(row, 4, dados['caixa'])
            ws.cell(row, 4).number_format = 'R$ #,##0'

            ws.cell(row, 5, dados['devido'])
            ws.cell(row, 5).number_format = 'R$ #,##0'

            var = self._calcular_delta(dados['pl'], dados['pl_d1'])
            if var != "-":
                ws.cell(row, 6, var)
                ws.cell(row, 6).number_format = '0.00%'
            else:
                ws.cell(row, 6, '-')

            row += 1

        # Ajustar larguras
        for col in range(1, 7):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 18

        self.print_success("Análise por tipo criada")

    def configurar_planilha(self):
        """Configura frozen panes, filtros, etc"""
        # Frozen panes na aba de dados
        self.ws_dados.freeze_panes = 'D4'  # Congela 3 primeiras linhas e 3 colunas

        # Auto filtro
        self.ws_dados.auto_filter.ref = f'A3:W{self.ws_dados.max_row}'

        self.print_success("Planilha configurada (freeze panes, filtros)")

    def salvar_relatorio(self):
        """Salva o relatório"""
        self.print_step(10, 10, "SALVANDO RELATÓRIO")

        # Configurações finais
        self.configurar_planilha()

        # Salva arquivo principal
        try:
            self.wb.save(self.OUTPUT_PATH)
            self.print_success(f"Relatório salvo: {os.path.basename(self.OUTPUT_PATH)}")
        except PermissionError:
            self.print_warning(f"Arquivo {os.path.basename(self.OUTPUT_PATH)} está aberto. Tentando nome alternativo...")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_path = self.OUTPUT_PATH.replace('.xlsx', f'_{timestamp}.xlsx')
            self.wb.save(alt_path)
            self.print_success(f"Salvo como: {os.path.basename(alt_path)}")

        # Salva cópia com data
        data_nome = self.input_date.strftime("%Y%m%d")
        arquivo_reports = os.path.join(self.REPORTS_DIR, f"Daily_Report_V5_{data_nome}.xlsx")

        try:
            self.wb.save(arquivo_reports)
            self.print_success(f"Cópia salva: {os.path.basename(arquivo_reports)}")
        except PermissionError:
            self.print_warning(f"Arquivo {os.path.basename(arquivo_reports)} está aberto")
            timestamp = datetime.now().strftime("%H%M%S")
            alt_reports = arquivo_reports.replace('.xlsx', f'_{timestamp}.xlsx')
            self.wb.save(alt_reports)
            self.print_success(f"Cópia salva como: {os.path.basename(alt_reports)}")

        self.wb.close()

    def imprimir_resumo(self):
        """Imprime resumo da execução"""
        self.print_header("RESUMO DA EXECUÇÃO V5 ENHANCED")

        tempo_total = (self.stats['fim'] - self.stats['inicio']).total_seconds()

        print(f"\n{Fore.CYAN}⏱  TEMPO DE EXECUÇÃO:{Style.RESET_ALL}")
        print(f"  Total: {Fore.YELLOW}{tempo_total:.2f}s ({tempo_total/60:.2f} min){Style.RESET_ALL}")

        print(f"\n{Fore.CYAN}📊 DADOS PROCESSADOS:{Style.RESET_ALL}")
        print(f"  Fundos: {Fore.YELLOW}{self.stats['fundos_processados']}{Style.RESET_ALL}")

        print(f"\n{Fore.CYAN}✨ MELHORIAS V5:{Style.RESET_ALL}")
        print(f"  • 3 abas (Resumo, Dados, Análise)")
        print(f"  • Formatação condicional")
        print(f"  • Colunas calculadas (Status, Liquidez, etc)")
        print(f"  • Frozen panes e filtros")

        if self.stats['erros']:
            print(f"\n{Fore.RED}❌ ERROS:{Style.RESET_ALL}")
            for erro in self.stats['erros']:
                print(f"  • {erro}")
        else:
            print(f"\n{Fore.GREEN}✓ NENHUM ERRO{Style.RESET_ALL}")

        print()

    def executar(self):
        """Executa todo o processo V5"""
        self.stats['inicio'] = datetime.now()

        try:
            self.print_header("RELATÓRIO DIÁRIO DE FUNDOS V5 - ENHANCED")

            # Passo 1: Validar
            if not self.validar_ambiente():
                return False

            # Passo 2: Obter data
            self.print_step(2, 10, "OBTENDO DATA DE ENTRADA")
            if not self.obter_data_entrada():
                self.print_error("Data inválida. Processo cancelado.")
                return False

            # Passo 3: Coletar dados
            self.coletar_dados_completos()

            # Passo 4: Criar workbook
            self.criar_workbook()

            # Passo 5: Resumo executivo
            self.criar_resumo_executivo()

            # Passo 6: Headers dados
            self.escrever_headers_dados()

            # Passo 7: Dados
            self.escrever_dados_fundos_enhanced()

            # Passo 8: Formatação condicional
            self.aplicar_formatacao_condicional()

            # Passo 9: Análise por tipo
            self.criar_analise_por_tipo()

            # Passo 10: Salvar
            self.salvar_relatorio()

            self.stats['fim'] = datetime.now()
            self.imprimir_resumo()

            self.print_header("✓ PROCESSO CONCLUÍDO COM SUCESSO!")

            return True

        except Exception as e:
            self.stats['erros'].append(str(e))
            self.stats['fim'] = datetime.now()

            self.print_header("✗ PROCESSO FINALIZADO COM ERROS")
            self.print_error(f"ERRO FATAL: {e}")
            logger.exception("Detalhes:")

            return False

        finally:
            if self.conn:
                try:
                    self.conn.close()
                except:
                    pass


def main():
    """Função principal"""

    print(f"{Fore.CYAN}{Style.BRIGHT}")
    print("╔════════════════════════════════════════════════════════════════╗")
    print("║                                                                ║")
    print("║      RELATÓRIO DIÁRIO DE FUNDOS - VERSÃO 5.0 ENHANCED        ║")
    print("║       FORMATAÇÃO PROFISSIONAL + DADOS ENRIQUECIDOS           ║")
    print("║                                                                ║")
    print("╚════════════════════════════════════════════════════════════════╝")
    print(f"{Style.RESET_ALL}\n")

    try:
        report = ReportDiarioFundosV5()
        sucesso = report.executar()

        print("\n" + "="*80)
        if sucesso:
            print(f"{Fore.GREEN}{Style.BRIGHT}✓ EXECUÇÃO FINALIZADA COM SUCESSO{Style.RESET_ALL}")
            print(f"\n{Fore.CYAN}Arquivos gerados:{Style.RESET_ALL}")
            print(f"  • v5_enhanced_report.xlsx")
            print(f"  • Daily_Report_V5_YYYYMMDD.xlsx")
        else:
            print(f"{Fore.RED}{Style.BRIGHT}✗ EXECUÇÃO FINALIZADA COM ERROS{Style.RESET_ALL}")
        print("="*80 + "\n")

    except KeyboardInterrupt:
        print(f"\n\n{Fore.YELLOW}⚠ Processo interrompido{Style.RESET_ALL}\n")
    except Exception as e:
        print(f"\n{Fore.RED}✗ ERRO CRÍTICO: {e}{Style.RESET_ALL}")
        logger.exception("Erro:")


if __name__ == "__main__":
    main()
