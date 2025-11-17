"""
REPORT DIÁRIO DE FUNDOS - Versão 4.0 (CONSTRUÇÃO DINÂMICA)
Gera o relatório completamente via Python, sem dependência de fórmulas Excel

Autor: Conversão para construção dinâmica
Data: 28/10/2025
Versão: 4.0 - SEM FÓRMULAS EXCEL

DIFERENÇAS V3 → V4:
- V3: Usa fórmulas XLOOKUP no Excel (colunas A-U)
- V4: Calcula TUDO em Python e escreve valores diretos
- V4: Não precisa de template com fórmulas
- V4: Gera arquivo do zero com formatação
"""

import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_PERCENTAGE_00, FORMAT_DATE_DDMMYY
from datetime import datetime
import os
import time
from typing import Dict, Any, Optional
import logging
from colorama import init, Fore, Back, Style

# Inicializa colorama
init(autoreset=True)

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('report_diario_v4.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ReportDiarioFundosV4:
    """Classe para geração DINÂMICA do relatório sem fórmulas Excel"""

    def __init__(self):
        """Inicializa caminhos e configurações"""
        # Caminhos
        self.DB_PATH = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\09. Base_de_Dados\Base Fundos_V2.accdb"
        self.OUTPUT_PATH = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\Report Diário\Teste V3 - Lucas\v4_temp_report.xlsx"
        self.FINAL_REPORT_PATH = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\Report Diário\Fundos - Report Diário.xlsx"
        self.ATUALIZAR_CARTEIRAS_PATH = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\Atualizar Carteiras_V8.xlsm"
        self.REPORTS_DIR = r"C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\Report Diário\REPORTS"

        # Workbook e worksheet
        self.wb = None
        self.ws = None
        self.conn = None

        # Data de entrada
        self.input_date = None
        self.input_date_str = None

        # Dicionário principal de dados por fundo
        self.fundos_dados = {}

        # Estatísticas
        self.stats = {
            'inicio': None,
            'fim': None,
            'fundos_processados': 0,
            'erros': []
        }

    # ===== MÉTODOS DE UTILIDADE =====

    def print_header(self, texto: str):
        """Imprime cabeçalho destacado"""
        largura = 80
        print("\n" + "=" * largura)
        print(Fore.CYAN + Style.BRIGHT + texto.center(largura))
        print("=" * largura + "\n")

    def print_step(self, numero: int, total: int, descricao: str):
        """Imprime passo da execução"""
        print(f"\n{Fore.YELLOW}[PASSO {numero}/{total}]{Style.RESET_ALL} {Fore.WHITE}{descricao}{Style.RESET_ALL}")
        print("-" * 80)

    def print_success(self, mensagem: str):
        """Imprime mensagem de sucesso"""
        print(f"{Fore.GREEN}✓{Style.RESET_ALL} {mensagem}")

    def print_warning(self, mensagem: str):
        """Imprime aviso"""
        print(f"{Fore.YELLOW}⚠{Style.RESET_ALL} {mensagem}")

    def print_error(self, mensagem: str):
        """Imprime erro"""
        print(f"{Fore.RED}✗{Style.RESET_ALL} {mensagem}")

    def print_info(self, mensagem: str):
        """Imprime informação"""
        print(f"{Fore.CYAN}ℹ{Style.RESET_ALL} {mensagem}")

    def validar_ambiente(self) -> bool:
        """Valida se todos os arquivos necessários existem"""
        self.print_step(1, 8, "VALIDANDO AMBIENTE")

        validacoes = {
            "Banco de dados Access": self.DB_PATH,
        }

        todas_validas = True
        for nome, caminho in validacoes.items():
            if os.path.exists(caminho):
                self.print_success(f"{nome}: OK")
            else:
                self.print_error(f"{nome}: NÃO ENCONTRADO")
                todas_validas = False

        # Cria diretórios se não existirem
        os.makedirs(os.path.dirname(self.OUTPUT_PATH), exist_ok=True)
        os.makedirs(self.REPORTS_DIR, exist_ok=True)
        self.print_success("Diretórios verificados/criados")

        return todas_validas

    def obter_data_entrada(self) -> Optional[datetime]:
        """Solicita a data de entrada ao usuário"""
        try:
            print(f"\n{Fore.CYAN}{'='*80}{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Por favor, informe a data de referência para o relatório:{Style.RESET_ALL}")
            print(f"{Fore.CYAN}Formatos aceitos: dd/mm/aaaa, dd-mm-aaaa, aaaa-mm-dd{Style.RESET_ALL}")
            print(f"{Fore.CYAN}Exemplo: 28/10/2025{Style.RESET_ALL}")
            print(f"{Fore.CYAN}{'='*80}{Style.RESET_ALL}\n")

            max_tentativas = 3
            tentativa = 0

            while tentativa < max_tentativas:
                data_input = input(f"{Fore.GREEN}Digite a data: {Style.RESET_ALL}").strip()

                if data_input.lower() in ('sair', 'exit', 'cancelar', 'q'):
                    self.print_warning("Operação cancelada pelo usuário")
                    return None

                if not data_input:
                    self.print_warning("Data não pode ser vazia")
                    tentativa += 1
                    continue

                parsed = None
                formatos = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%d/%Y"]

                for fmt in formatos:
                    try:
                        parsed = datetime.strptime(data_input, fmt)
                        break
                    except ValueError:
                        continue

                if parsed:
                    parsed = parsed.replace(hour=0, minute=0, second=0, microsecond=0)
                    self.input_date = parsed
                    self.input_date_str = parsed.strftime("%m/%d/%Y")

                    print()
                    self.print_success(f"Data aceita: {Fore.YELLOW}{parsed.strftime('%d/%m/%Y')}{Style.RESET_ALL}")
                    self.print_info(f"Formato Access: {self.input_date_str}")
                    print()

                    return self.input_date
                else:
                    tentativa += 1
                    self.print_error(f"Data inválida: {data_input}")
                    if tentativa < max_tentativas:
                        self.print_warning(f"Tentativa {tentativa}/{max_tentativas}. Tente novamente.")
                    else:
                        self.print_error("Número máximo de tentativas excedido")

            return None

        except KeyboardInterrupt:
            print(f"\n\n{Fore.YELLOW}⚠ Entrada cancelada pelo usuário{Style.RESET_ALL}\n")
            return None
        except Exception as e:
            self.print_error(f"Falha ao obter data: {e}")
            return None

    def conectar_access(self) -> pyodbc.Connection:
        """Conecta ao banco Access"""
        try:
            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={self.DB_PATH};'
            )
            conn = pyodbc.connect(conn_str)
            self.print_success("Conexão com banco Access estabelecida")
            return conn
        except Exception as e:
            self.print_error(f"Erro ao conectar ao Access: {e}")
            raise

    # ===== COLETA DE DADOS =====

    def coletar_dados_completos(self):
        """Coleta TODOS os dados necessários do banco em uma única passada"""
        self.print_step(3, 8, "COLETANDO DADOS DO BANCO ACCESS")

        self.conn = self.conectar_access()

        # Query principal UNION (mesma do V2)
        self.print_info("Executando query principal...")
        df_principal = self._executar_query_principal()

        # Processar cada fundo da query principal
        for _, row in df_principal.iterrows():
            fundo = row['CARTEIRA']
            if fundo not in self.fundos_dados:
                self.fundos_dados[fundo] = {}

            self.fundos_dados[fundo].update({
                'data_base': row['DATA_BASE'],
                'pl_posicao_ativos': row['PL_Posicao_Ativos'],
                'pl': row['PL'],
                'pl_d1': row['PL_D1'],
                'pl_d7': row['PL_D7'],
                'pl_d30': row['PL_D30'],
                'caixa_bancario': row['SALDOCAIXAATUAL'],
                'fundos_di': row['CAIXA_REAG_II'],
                'devido_taxas': row['devido_taxa'],
                'caixa_total': row['CAIXA_TOTAL'],
                'cotista': row.get('cotista'),
                'vl_bruto': row.get('vl_bruto'),
                'vl_liquido': row.get('vl_liquido'),
                'tipo': self._determinar_tipo(fundo)
            })

        self.print_success(f"Query principal: {len(self.fundos_dados)} fundos")

        # Processar 19 FIDCs específicos
        self._coletar_dados_fidcs()

        # Processar 24 fundos Qore
        self._coletar_dados_qore()

        self.print_success(f"Total de fundos coletados: {len(self.fundos_dados)}")
        self.stats['fundos_processados'] = len(self.fundos_dados)

        if self.conn:
            self.conn.close()
            self.conn = None

    def _executar_query_principal(self) -> pd.DataFrame:
        """Executa a query SQL principal (UNION de 2 blocos)"""
        query = f"""
        SELECT PT.DATA_BASE, PT.CARTEIRA, PT.VALORPOSICAOATIVOS AS PL_Posicao_Ativos,
               PT.VALORPATRIMONIOLIQUIDO AS PL,
               (SELECT TOP 1 PT1.VALORPATRIMONIOLIQUIDO FROM Patrimonio_Totais AS PT1
                WHERE PT1.CARTEIRA = PT.CARTEIRA AND PT1.DATA_BASE =
                (SELECT DU.[D-1] FROM Dias_Uteis AS DU WHERE DU.Data = PT.DATA_BASE)) AS PL_D1,
               (SELECT TOP 1 PT7.VALORPATRIMONIOLIQUIDO FROM Patrimonio_Totais AS PT7
                WHERE PT7.CARTEIRA = PT.CARTEIRA AND PT7.DATA_BASE =
                (SELECT DU.[D-7] FROM Dias_Uteis AS DU WHERE DU.Data = PT.DATA_BASE)) AS PL_D7,
               (SELECT TOP 1 PT30.VALORPATRIMONIOLIQUIDO FROM Patrimonio_Totais AS PT30
                WHERE PT30.CARTEIRA = PT.CARTEIRA AND PT30.DATA_BASE =
                (SELECT DU.[D-30] FROM Dias_Uteis AS DU WHERE DU.Data = PT.DATA_BASE)) AS PL_D30,
               PT.SALDOCAIXAATUAL,
               IIF(ISNULL((SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                           WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                           AND FUNDS.CARTEIRA = PT.CARTEIRA)), 0,
                   (SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                    WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                    AND FUNDS.CARTEIRA = PT.CARTEIRA)) AS CAIXA_REAG_II,
               (SELECT SUM(VALOR) FROM CPR_Lancamentos AS CPR WHERE CPR.DESCRICAO LIKE 'taxa%'
                AND CPR.CARTEIRA = PT.CARTEIRA AND CPR.DATA_BASE = PT.DATA_BASE) AS devido_taxa,
               IIF(ISNULL(PT.SALDOCAIXAATUAL), 0, PT.SALDOCAIXAATUAL) +
               IIF(ISNULL((SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                           WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                           AND FUNDS.CARTEIRA = PT.CARTEIRA)), 0,
                   (SELECT TOP 1 FUNDS.VALORLIQUIDO FROM Fundos_Fundos AS FUNDS
                    WHERE FUNDS.CODIGO = 'FIRF REAG CA II' AND FUNDS.DATA_BASE = PT.DATA_BASE
                    AND FUNDS.CARTEIRA = PT.CARTEIRA)) AS CAIXA_TOTAL,
               jc.cotista, jc.vl_bruto, jc.vl_liquido
        FROM Patrimonio_Totais AS PT
        LEFT JOIN (SELECT cd_fundo, DT_POSICAO AS data_base, MAX(CD_COTISTA) AS COTISTA,
                          SUM(VL_CORRIGIDO) AS vl_bruto, SUM(VL_RESGATE) AS vl_liquido
                   FROM jcot GROUP BY cd_fundo, DT_POSICAO) AS jc
        ON (jc.data_base = PT.DATA_BASE) AND (jc.cd_fundo = PT.CARTEIRA)
        WHERE PT.DATA_BASE = #{self.input_date_str}#

        UNION ALL

        SELECT CP.[Data Posição] AS DATA_BASE, CP.FUNDO AS CARTEIRA,
               (SELECT TOP 1 RV.ValorTotal FROM Renda_Variável_Maps AS RV
                WHERE RV.FUNDO = CP.FUNDO AND RV.DATA_INPUT = CP.[Data Posição]) AS PL_Posicao_Ativos,
               CP.[PL Posição] AS PL,
               (SELECT TOP 1 CP1.[PL Posição] FROM Cotas_Patrimonio_Maps AS CP1
                WHERE CP1.FUNDO = CP.FUNDO AND CP1.[Data Posição] =
                (SELECT DU.[D-1] FROM Dias_Uteis AS DU WHERE DU.Data = CP.[Data Posição])) AS PL_D1,
               (SELECT TOP 1 CP7.[PL Posição] FROM Cotas_Patrimonio_Maps AS CP7
                WHERE CP7.FUNDO = CP.FUNDO AND CP7.[Data Posição] =
                (SELECT DU.[D-7] FROM Dias_Uteis AS DU WHERE DU.Data = CP.[Data Posição])) AS PL_D7,
               (SELECT TOP 1 CP30.[PL Posição] FROM Cotas_Patrimonio_Maps AS CP30
                WHERE CP30.FUNDO = CP.FUNDO AND CP30.[Data Posição] =
                (SELECT DU.[D-30] FROM Dias_Uteis AS DU WHERE DU.Data = CP.[Data Posição])) AS PL_D30,
               (SELECT TOP 1 CM.ValorTotal FROM Caixa_Maps AS CM
                WHERE CM.FUNDO = CP.FUNDO AND CM.DATA_INPUT = CP.[Data Posição]) AS SALDOCAIXAATUAL,
               IIF(ISNULL((SELECT TOP 1 CFM.ValorTotal FROM Cotas_Fundos_Maps AS CFM
                           WHERE CFM.FUNDO = CP.FUNDO AND CFM.DATA_INPUT = CP.[Data Posição])), 0,
                   (SELECT TOP 1 CFM.ValorTotal FROM Cotas_Fundos_Maps AS CFM
                    WHERE CFM.FUNDO = CP.FUNDO AND CFM.DATA_INPUT = CP.[Data Posição])) AS CAIXA_REAG_II,
               (SELECT SUM(VP.ValorTotal) FROM Valor_a_Pagar_Maps AS VP
                WHERE VP.FUNDO = CP.FUNDO AND VP.DATA_INPUT = CP.[Data Posição]) AS devido_taxa,
               NULL AS CAIXA_TOTAL, NULL AS cotista,
               IIF(CP.FUNDO IN ('FIP TRAIL MULT','FIP GEKKO','FIP RAM','FIP BENELLI',
                                'FIP TURANO','FIP MURCIELAGO','FIP AMG','FIP OSLO','FIP ESTOCOLMO'),
                   IIF(ISNULL((SELECT SUM(P.VLIR) FROM Passivo_Maps AS P
                               WHERE MID(P.FUNDO, 7) = CP.FUNDO AND P.[DATA_INPUT] = CP.[Data Posição])), 0,
                       (SELECT SUM(P.VLIR) FROM Passivo_Maps AS P
                        WHERE MID(P.FUNDO, 7) = CP.FUNDO AND P.[DATA_INPUT] = CP.[Data Posição])),
                   NULL) AS vl_bruto,
               NULL AS vl_liquido
        FROM Cotas_Patrimonio_Maps AS CP
        WHERE CP.[Data Posição] = #{self.input_date_str}#
        AND CP.FUNDO IN ('FIP TRAIL MULT', 'FIP GEKKO', 'FIP RAM', 'FIP BENELLI', 'FIP TURANO',
                         'FIP MURCIELAGO', 'FIP AMG', 'FIP OSLO', 'FIP ESTOCOLMO')
        """

        inicio = time.time()
        df = pd.read_sql(query, self.conn)
        tempo = time.time() - inicio

        self.print_info(f"Query executada em {tempo:.2f}s - {len(df)} registros")
        return df

    def _determinar_tipo(self, fundo: str) -> str:
        """Determina o tipo do fundo baseado no nome"""
        fundo_upper = fundo.upper()
        if 'FIDC' in fundo_upper:
            return 'FIDC'
        elif 'FIP' in fundo_upper:
            return 'FIP'
        elif 'FIM' in fundo_upper:
            return 'FIM'
        else:
            return '-'

    def _coletar_dados_fidcs(self):
        """Coleta dados dos 19 FIDCs do sistema MAPS"""
        self.print_info("Coletando dados FIDCs...")

        fundos_fidcs = [
            "FIDC FANGIO", "FIDC FLYNN", "FIDC HILL NP", "FIDC IRIDIO", "FIDC KOLEOS",
            "FIDC LEGACY", "FIDC LOGAN", "FIDC MANSELL", "FIDC NORFOLK", "FIDC RAIKKONEN",
            "FIDC RINDT", "FIDC STARK", "FIDC TERPSICORE", "FIDC URANO NP", "FIDC VALKYRIES",
            "FIDC VANTAGE", "FIDC VIRAGE", "FIDC WILDE NP", "FIDC ZANDVOORT NP"
        ]

        in_list = "', '".join(fundos_fidcs)
        in_clause = f"('{in_list}')"

        # Obter datas úteis
        query_du = f"SELECT [D-1] AS D1, [D-7] AS D7, [D-30] AS D30 FROM Dias_Uteis WHERE Data = #{self.input_date_str}#"
        df_du = pd.read_sql(query_du, self.conn)

        if df_du.empty:
            self.print_warning("Dias úteis não encontrados para FIDCs")
            return

        d1_str = df_du['D1'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D1'].iloc[0]) else None
        d7_str = df_du['D7'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D7'].iloc[0]) else None
        d30_str = df_du['D30'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D30'].iloc[0]) else None

        # Queries de dados
        queries = {
            'Caixa': f"SELECT FUNDO, SUM(ValorTotal) AS SomaCaixa FROM Caixa_Maps WHERE DATA_INPUT = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO",
            'PL': f"SELECT FUNDO, MAX([PL Posição]) AS PL_POSICAO FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO",
            'VAP': f"SELECT FUNDO, SUM(ValorTotal) AS SomaValorTotal FROM Valor_a_Pagar_Maps WHERE DATA_INPUT = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO",
            'Renda Fixa': f"SELECT FUNDO, SUM(Valor_Total) AS SomaRF FROM Renda_Fixa_Maps WHERE DATA_INPUT = #{self.input_date_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO"
        }

        dados = {}
        for nome, query in queries.items():
            df = pd.read_sql(query, self.conn)
            dados[nome] = df

        # PL histórico
        if d1_str:
            dados['PL_D1'] = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D1 FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{d1_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO", self.conn)
        if d7_str:
            dados['PL_D7'] = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D7 FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{d7_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO", self.conn)
        if d30_str:
            dados['PL_D30'] = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D30 FROM Cotas_Patrimonio_Maps WHERE [Data Posição] = #{d30_str}# AND FUNDO IN {in_clause} GROUP BY FUNDO", self.conn)

        # Processar cada FIDC
        fidcs_encontrados = 0
        for fundo in fundos_fidcs:
            # Buscar valores
            rf_val = dados['Renda Fixa'][dados['Renda Fixa']['FUNDO'] == fundo]['SomaRF'].values
            pl_val = dados['PL'][dados['PL']['FUNDO'] == fundo]['PL_POSICAO'].values
            caixa_val = dados['Caixa'][dados['Caixa']['FUNDO'] == fundo]['SomaCaixa'].values
            vap_val = dados['VAP'][dados['VAP']['FUNDO'] == fundo]['SomaValorTotal'].values

            pl_d1_val = dados.get('PL_D1', pd.DataFrame())[dados.get('PL_D1', pd.DataFrame())['FUNDO'] == fundo]['PL_D1'].values if 'PL_D1' in dados else []
            pl_d7_val = dados.get('PL_D7', pd.DataFrame())[dados.get('PL_D7', pd.DataFrame())['FUNDO'] == fundo]['PL_D7'].values if 'PL_D7' in dados else []
            pl_d30_val = dados.get('PL_D30', pd.DataFrame())[dados.get('PL_D30', pd.DataFrame())['FUNDO'] == fundo]['PL_D30'].values if 'PL_D30' in dados else []

            pl_posicao_ativos = rf_val[0] if len(rf_val) > 0 else 0
            pl = pl_val[0] if len(pl_val) > 0 else 0
            caixa_bancario = caixa_val[0] if len(caixa_val) > 0 else 0
            devido_taxas = vap_val[0] if len(vap_val) > 0 else 0

            # Adicionar ao dicionário principal
            self.fundos_dados[fundo] = {
                'data_base': self.input_date,
                'pl_posicao_ativos': pl_posicao_ativos,
                'pl': pl,
                'pl_d1': pl_d1_val[0] if len(pl_d1_val) > 0 else 0,
                'pl_d7': pl_d7_val[0] if len(pl_d7_val) > 0 else 0,
                'pl_d30': pl_d30_val[0] if len(pl_d30_val) > 0 else 0,
                'caixa_bancario': caixa_bancario,
                'fundos_di': 0,
                'devido_taxas': devido_taxas,
                'caixa_total': caixa_bancario,
                'cotista': None,
                'vl_bruto': 0,
                'vl_liquido': 0,
                'tipo': 'FIDC'
            }

            if pl > 0:
                fidcs_encontrados += 1

        self.print_success(f"FIDCs: {fidcs_encontrados}/{len(fundos_fidcs)} com dados")

    def _chave_fundo_padrao(self, nome: str) -> str:
        """Padroniza nome do fundo"""
        k = nome.upper().strip()
        k = ' '.join(k.split())

        # Casos específicos (mais específico primeiro)
        if 'BLOKO URBANISMO' in k or 'FIP BLOKO URBANISMO' in k:
            return 'FIP BLOKO URBANISMO'
        elif 'BLOKO FIM' in k or 'FIM BLOKO' in k:
            return 'FIM BLOKO'
        elif 'FIP GTB' in k:
            return 'FIP GTB MULT'
        elif 'FIP BLOKO' in k:
            return 'FIP BLOKO MULT'
        else:
            return k

    def _coletar_dados_qore(self):
        """Coleta dados dos 24 fundos do sistema Qore"""
        self.print_info("Coletando dados Qore...")

        fundos_qore = [
            "FIP MINAS", "FIP GOIAS", "FIP TERRAVISTA", "FIP BLOKO URBANISMO",
            "FIP CALDAS", "FIP KAWANA", "FIP RENOGRID", "FIP GTB MULT",
            "FIP GEKKO", "FIP RAM", "FIP ON", "FIP BENELLI",
            "FIP TURANO", "FIP MURCIELAGO", "FIP AMG", "FIP OSLO", "FIP ESTOCOLMO",
            "FIDC FORSETI", "FIDC SOCRATES", "FIDC PLATAO", "FIDC EVOQUE",
            "FIM PES", "FIDC SOA", "FIM BLOKO"
        ]

        in_list = "', '".join(fundos_qore)
        in_clause = f"('{in_list}')"
        extra_like = " OR FUNDO LIKE '*FIP GTB*' OR FUNDO LIKE '*BLOKO*'"

        # Datas úteis
        query_du = f"SELECT [D-1] AS D1, [D-7] AS D7, [D-30] AS D30 FROM Dias_Uteis WHERE Data = #{self.input_date_str}#"
        df_du = pd.read_sql(query_du, self.conn)

        if df_du.empty:
            self.print_warning("Dias úteis não encontrados para Qore")
            return

        d1_str = df_du['D1'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D1'].iloc[0]) else None
        d7_str = df_du['D7'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D7'].iloc[0]) else None
        d30_str = df_du['D30'].iloc[0].strftime("%m/%d/%Y") if pd.notna(df_du['D30'].iloc[0]) else None

        # Dicionários para agregar
        mapas = {
            'caixa': {}, 'pl': {}, 'pl_d1': {}, 'pl_d7': {}, 'pl_d30': {},
            'taxa': {}, 'pl_ativos': {}, 'reag_ii': {}
        }

        # 1. Caixa_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(Valor) AS ValorDia FROM Caixa_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['caixa'][key] = mapas['caixa'].get(key, 0) + (row['ValorDia'] if pd.notna(row['ValorDia']) else 0)

        # 2. CPR_QORE (taxa)
        try:
            df = pd.read_sql(f"SELECT FUNDO, SUM(VALOR) AS SomaTaxa FROM CPR_QORE WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) AND (Descrição LIKE 'taxa%' OR Descrição LIKE 'despesa%') GROUP BY FUNDO", self.conn)
        except:
            df = pd.read_sql(f"SELECT FUNDO, SUM(VALOR) AS SomaTaxa FROM CPR_QORE WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) AND (DESCRICAO LIKE 'taxa%' OR DESCRICAO LIKE 'despesa%') GROUP BY FUNDO", self.conn)

        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['taxa'][key] = mapas['taxa'].get(key, 0) + (row['SomaTaxa'] if pd.notna(row['SomaTaxa']) else 0)

        # 3. Sociedade_Limitada_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(IIF([Valor_Mercado] IS NULL OR [Valor_Mercado]=0, [Valor_Custo], [Valor_Mercado])) AS PL_Posicao_Ativos FROM Sociedade_Limitada_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl_ativos'][key] = mapas['pl_ativos'].get(key, 0) + (row['PL_Posicao_Ativos'] if pd.notna(row['PL_Posicao_Ativos']) else 0)

        # 4. Renda_Fixa_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(Valor_Bruto) AS PL_Posicao_Ativos FROM Renda_Fixa_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl_ativos'][key] = mapas['pl_ativos'].get(key, 0) + (row['PL_Posicao_Ativos'] if pd.notna(row['PL_Posicao_Ativos']) else 0)

        # 5. Direito_Creditorio_Qore
        df = pd.read_sql(f"SELECT FUNDO, SUM(IIF([Valor_Mercado] IS NULL OR [Valor_Mercado]=0, [Valor_Custo], [Valor_Mercado])) AS PL_Posicao_Ativos FROM Direito_Creditorio_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl_ativos'][key] = mapas['pl_ativos'].get(key, 0) + (row['PL_Posicao_Ativos'] if pd.notna(row['PL_Posicao_Ativos']) else 0)

        # 6. REAG II
        df = pd.read_sql(f"SELECT FUNDO, SUM([Qtde]*[Pu_Mercado]) AS ValorReagII FROM Sociedade_Limitada_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) AND [Descricao] LIKE '*REAG FIRF II*' GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['reag_ii'][key] = mapas['reag_ii'].get(key, 0) + (row['ValorReagII'] if pd.notna(row['ValorReagII']) else 0)

        # 7. PL atual
        df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_POSICAO FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{self.input_date_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
        for _, row in df.iterrows():
            key = self._chave_fundo_padrao(row['FUNDO'])
            mapas['pl'][key] = row['PL_POSICAO'] if pd.notna(row['PL_POSICAO']) else 0

        # 8-10. PL histórico
        if d1_str:
            df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D1 FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{d1_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
            for _, row in df.iterrows():
                key = self._chave_fundo_padrao(row['FUNDO'])
                mapas['pl_d1'][key] = row['PL_D1'] if pd.notna(row['PL_D1']) else 0

        if d7_str:
            df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D7 FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{d7_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
            for _, row in df.iterrows():
                key = self._chave_fundo_padrao(row['FUNDO'])
                mapas['pl_d7'][key] = row['PL_D7'] if pd.notna(row['PL_D7']) else 0

        if d30_str:
            df = pd.read_sql(f"SELECT FUNDO, MAX([PL Posição]) AS PL_D30 FROM Cotas_Patrimonio_Qore WHERE DATA_INPUT = #{d30_str}# AND (FUNDO IN {in_clause} {extra_like}) GROUP BY FUNDO", self.conn)
            for _, row in df.iterrows():
                key = self._chave_fundo_padrao(row['FUNDO'])
                mapas['pl_d30'][key] = row['PL_D30'] if pd.notna(row['PL_D30']) else 0

        # Processar cada fundo Qore
        qore_encontrados = 0
        for fundo in fundos_qore:
            key = self._chave_fundo_padrao(fundo)

            caixa = mapas['caixa'].get(key, 0)
            fundos_di = mapas['reag_ii'].get(key, 0)
            pl = mapas['pl'].get(key, 0)

            self.fundos_dados[key] = {
                'data_base': self.input_date,
                'pl_posicao_ativos': mapas['pl_ativos'].get(key, 0),
                'pl': pl,
                'pl_d1': mapas['pl_d1'].get(key, 0),
                'pl_d7': mapas['pl_d7'].get(key, 0),
                'pl_d30': mapas['pl_d30'].get(key, 0),
                'caixa_bancario': caixa,
                'fundos_di': fundos_di,
                'devido_taxas': mapas['taxa'].get(key, 0),
                'caixa_total': caixa + fundos_di,
                'cotista': None,
                'vl_bruto': 0,
                'vl_liquido': 0,
                'tipo': self._determinar_tipo(key)
            }

            if pl > 0:
                qore_encontrados += 1

        self.print_success(f"Qore: {qore_encontrados}/{len(fundos_qore)} com dados")

    # ===== CONSTRUÇÃO DO RELATÓRIO =====

    def criar_workbook(self):
        """Cria novo workbook do zero"""
        self.print_step(4, 8, "CRIANDO ESTRUTURA DO RELATÓRIO")

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Output"

        self.print_success("Workbook criado")

    def escrever_headers(self):
        """Escreve headers do relatório (linhas 1 e 2)"""
        self.print_step(5, 8, "ESCREVENDO HEADERS")

        # Linha 1 - Categorias principais
        headers_l1 = {
            5: "Patrimônio Líquido",
            9: "Caixa",
            14: "Passivo",
            16: "Patrimônio Líquido"
        }

        for col, texto in headers_l1.items():
            cell = self.ws.cell(1, col, texto)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Linha 2 - Headers detalhados
        headers_l2 = [
            "Fundo", "Gestora", "Operação", "Posição Ativos",
            "Atual", "∆ D-1", "∆ D-7", "∆ D-30",
            "Saldo Bancário", "Fundos DI", "Total", "Devido Taxas", "NCG",
            "Importo Retido", "",
            "D-1", "D-7", "D-30",
            "", "", "TIPO"
        ]

        for col, texto in enumerate(headers_l2, start=1):
            cell = self.ws.cell(2, col, texto)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            # Bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

        # Ajustar larguras das colunas
        larguras = {
            'A': 25, 'B': 15, 'C': 12, 'D': 15,
            'E': 15, 'F': 10, 'G': 10, 'H': 10,
            'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 12,
            'N': 15, 'O': 5,
            'P': 15, 'Q': 15, 'R': 15,
            'S': 5, 'T': 5, 'U': 10
        }

        for col_letter, width in larguras.items():
            self.ws.column_dimensions[col_letter].width = width

        self.print_success("Headers escritos")

    def escrever_dados_fundos(self):
        """Escreve dados de todos os fundos com CÁLCULOS DINÂMICOS"""
        self.print_step(6, 8, "ESCREVENDO DADOS DOS FUNDOS")

        # Ordena fundos por nome
        fundos_ordenados = sorted(self.fundos_dados.keys())

        row = 3  # Começa na linha 3
        fundos_escritos = 0

        for fundo in fundos_ordenados:
            dados = self.fundos_dados[fundo]

            # Coluna A: Fundo
            self.ws.cell(row, 1, fundo)

            # Coluna B: Gestora (buscar de alguma tabela ou deixar em branco)
            self.ws.cell(row, 2, self._buscar_gestora(fundo))

            # Coluna C: Operação
            self.ws.cell(row, 3, "-")

            # Coluna D: Posição Ativos (PL_Posicao_Ativos - Fundos_DI)
            pl_pos_ativos = dados.get('pl_posicao_ativos', 0) or 0
            fundos_di = dados.get('fundos_di', 0) or 0
            self.ws.cell(row, 4, pl_pos_ativos - fundos_di)

            # Coluna E: PL Atual
            pl_atual = dados.get('pl', 0) or 0
            self.ws.cell(row, 5, pl_atual)

            # Colunas F, G, H: Delta D-1, D-7, D-30
            pl_d1 = dados.get('pl_d1', 0) or 0
            pl_d7 = dados.get('pl_d7', 0) or 0
            pl_d30 = dados.get('pl_d30', 0) or 0

            self.ws.cell(row, 6, self._calcular_delta(pl_atual, pl_d1))
            self.ws.cell(row, 7, self._calcular_delta(pl_atual, pl_d7))
            self.ws.cell(row, 8, self._calcular_delta(pl_atual, pl_d30))

            # Coluna I: Saldo Bancário
            caixa_bancario = dados.get('caixa_bancario', 0) or 0
            self.ws.cell(row, 9, caixa_bancario)

            # Coluna J: Fundos DI
            self.ws.cell(row, 10, fundos_di)

            # Coluna K: Total Caixa
            total_caixa = dados.get('caixa_total', 0) or caixa_bancario + fundos_di
            self.ws.cell(row, 11, total_caixa)

            # Coluna L: Devido Taxas
            devido_taxas = dados.get('devido_taxas', 0) or 0
            self.ws.cell(row, 12, devido_taxas)

            # Coluna M: NCG (Necessidade de Capital de Giro)
            ncg = self._calcular_ncg(devido_taxas, total_caixa)
            self.ws.cell(row, 13, ncg)

            # Coluna N: Importo Retido (vl_bruto - vl_liquido)
            vl_bruto = dados.get('vl_bruto', 0) or 0
            vl_liquido = dados.get('vl_liquido', 0) or 0
            self.ws.cell(row, 14, vl_bruto - vl_liquido)

            # Coluna O: vazio

            # Colunas P, Q, R: PL histórico
            self.ws.cell(row, 16, pl_d1)
            self.ws.cell(row, 17, pl_d7)
            self.ws.cell(row, 18, pl_d30)

            # Colunas S, T: vazias

            # Coluna U: TIPO
            self.ws.cell(row, 21, dados.get('tipo', '-'))

            row += 1
            fundos_escritos += 1

        self.print_success(f"Dados escritos: {fundos_escritos} fundos")

    def _buscar_gestora(self, fundo: str) -> str:
        """Busca gestora do fundo (implementar se necessário)"""
        # Mapeamento simples - expandir conforme necessário
        mapeamento = {
            'FIDC FORSETI': 'North Sea',
            'FIDC SOCRATES': 'North Sea',
            'FIDC PLATAO': 'North Sea',
            # Adicionar mais conforme necessário
        }
        return mapeamento.get(fundo, '-')

    def _calcular_delta(self, valor_atual, valor_anterior) -> float:
        """Calcula variação percentual"""
        if not valor_anterior or valor_anterior == 0:
            return "-"

        if valor_anterior < 0:
            if valor_atual > 0:
                return "-"
            else:
                return round(1 - valor_atual / valor_anterior, 3)
        else:
            if valor_atual < 0:
                return "-"
            else:
                return round(valor_atual / valor_anterior - 1, 3)

    def _calcular_ncg(self, devido_taxas, caixa_total) -> float:
        """Calcula Necessidade de Capital de Giro"""
        if -devido_taxas > caixa_total:
            return -(devido_taxas + caixa_total)
        return 0

    def aplicar_formatacao(self):
        """Aplica formatação numérica e visual"""
        self.print_step(7, 8, "APLICANDO FORMATAÇÃO")

        # Formatar colunas numéricas
        for row in range(3, self.ws.max_row + 1):
            # Colunas D, E, I, J, K, L, M, N, P, Q, R: formato numérico com separador de milhares
            for col in [4, 5, 9, 10, 11, 12, 13, 14, 16, 17, 18]:
                cell = self.ws.cell(row, col)
                if isinstance(cell.value, (int, float)) and cell.value != "-":
                    cell.number_format = '#,##0'

            # Colunas F, G, H: formato percentual
            for col in [6, 7, 8]:
                cell = self.ws.cell(row, col)
                if isinstance(cell.value, (int, float)) and cell.value != "-":
                    cell.number_format = '0.00%'

        self.print_success("Formatação aplicada")

    def salvar_relatorio(self):
        """Salva o relatório"""
        self.print_step(8, 8, "SALVANDO RELATÓRIO")

        # Salva arquivo principal
        try:
            self.wb.save(self.OUTPUT_PATH)
            self.print_success(f"Relatório salvo: {os.path.basename(self.OUTPUT_PATH)}")
        except PermissionError:
            self.print_warning(f"Arquivo {os.path.basename(self.OUTPUT_PATH)} está aberto. Tentando nome alternativo...")
            # Tenta salvar com timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_path = self.OUTPUT_PATH.replace('.xlsx', f'_{timestamp}.xlsx')
            self.wb.save(alt_path)
            self.print_success(f"Salvo como: {os.path.basename(alt_path)}")

        # Salva cópia com data no REPORTS
        data_nome = self.input_date.strftime("%Y%m%d")
        arquivo_reports = os.path.join(self.REPORTS_DIR, f"Daily_Report_V4_{data_nome}.xlsx")

        try:
            self.wb.save(arquivo_reports)
            self.print_success(f"Cópia salva: {os.path.basename(arquivo_reports)}")
        except PermissionError:
            self.print_warning(f"Arquivo {os.path.basename(arquivo_reports)} está aberto")
            # Tenta com timestamp
            timestamp = datetime.now().strftime("%H%M%S")
            alt_reports = arquivo_reports.replace('.xlsx', f'_{timestamp}.xlsx')
            self.wb.save(alt_reports)
            self.print_success(f"Cópia salva como: {os.path.basename(alt_reports)}")

        self.wb.close()

    def imprimir_resumo(self):
        """Imprime resumo da execução"""
        self.print_header("RESUMO DA EXECUÇÃO")

        tempo_total = (self.stats['fim'] - self.stats['inicio']).total_seconds()

        print(f"\n{Fore.CYAN}⏱  TEMPO DE EXECUÇÃO:{Style.RESET_ALL}")
        print(f"  Total: {Fore.YELLOW}{tempo_total:.2f}s ({tempo_total/60:.2f} min){Style.RESET_ALL}")

        print(f"\n{Fore.CYAN}📊 DADOS PROCESSADOS:{Style.RESET_ALL}")
        print(f"  Fundos: {Fore.YELLOW}{self.stats['fundos_processados']}{Style.RESET_ALL}")

        if self.stats['erros']:
            print(f"\n{Fore.RED}❌ ERROS:{Style.RESET_ALL}")
            for erro in self.stats['erros']:
                print(f"  • {erro}")
        else:
            print(f"\n{Fore.GREEN}✓ NENHUM ERRO{Style.RESET_ALL}")

        print()

    def executar(self):
        """Executa todo o processo de geração do relatório V4"""
        self.stats['inicio'] = datetime.now()

        try:
            self.print_header("RELATÓRIO DIÁRIO DE FUNDOS V4 - CONSTRUÇÃO DINÂMICA")

            # Passo 1: Validar
            if not self.validar_ambiente():
                return False

            # Passo 2: Obter data
            self.print_step(2, 8, "OBTENDO DATA DE ENTRADA")
            if not self.obter_data_entrada():
                self.print_error("Data inválida. Processo cancelado.")
                return False

            # Passo 3: Coletar dados
            self.coletar_dados_completos()

            # Passo 4: Criar workbook
            self.criar_workbook()

            # Passo 5: Headers
            self.escrever_headers()

            # Passo 6: Dados
            self.escrever_dados_fundos()

            # Passo 7: Formatação
            self.aplicar_formatacao()

            # Passo 8: Salvar
            self.salvar_relatorio()

            self.stats['fim'] = datetime.now()
            self.imprimir_resumo()

            self.print_header("✓ PROCESSO CONCLUÍDO COM SUCESSO!")

            return True

        except Exception as e:
            self.stats['erros'].append(str(e))
            self.stats['fim'] = datetime.now()

            self.print_header("✗ PROCESSO FINALIZADO COM ERROS")
            self.print_error(f"ERRO FATAL: {e}")
            logger.exception("Detalhes:")

            return False

        finally:
            if self.conn:
                try:
                    self.conn.close()
                except:
                    pass  # Conexão já fechada


def main():
    """Função principal"""

    print(f"{Fore.CYAN}{Style.BRIGHT}")
    print("╔════════════════════════════════════════════════════════════════╗")
    print("║                                                                ║")
    print("║      RELATÓRIO DIÁRIO DE FUNDOS - VERSÃO 4.0 DINÂMICA          ║")
    print("║              SEM FÓRMULAS EXCEL - 100% PYTHON                  ║")
    print("║                                                                ║")
    print("╚════════════════════════════════════════════════════════════════╝")
    print(f"{Style.RESET_ALL}\n")

    try:
        report = ReportDiarioFundosV4()
        sucesso = report.executar()

        print("\n" + "="*80)
        if sucesso:
            print(f"{Fore.GREEN}{Style.BRIGHT}✓ EXECUÇÃO FINALIZADA COM SUCESSO{Style.RESET_ALL}")
            print(f"\n{Fore.CYAN}Arquivos gerados:{Style.RESET_ALL}")
            print(f"  • v4_temp_report.xlsx")
            print(f"  • Daily_Report_V4_YYYYMMDD.xlsx")
        else:
            print(f"{Fore.RED}{Style.BRIGHT}✗ EXECUÇÃO FINALIZADA COM ERROS{Style.RESET_ALL}")
        print("="*80 + "\n")

    except KeyboardInterrupt:
        print(f"\n\n{Fore.YELLOW}⚠ Processo interrompido{Style.RESET_ALL}\n")
    except Exception as e:
        print(f"\n{Fore.RED}✗ ERRO CRÍTICO: {e}{Style.RESET_ALL}")
        logger.exception("Erro:")


if __name__ == "__main__":
    main()
