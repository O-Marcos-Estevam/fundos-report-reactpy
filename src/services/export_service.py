"""
Serviço de Exportação
Exporta dados do dashboard para Excel, CSV e PDF
"""

import io
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.fundo import FundoData
from utils.analytics import calcular_metricas_fundo, calcular_health_score


class ExportService:
    """Serviço centralizado de exportação"""

    @staticmethod
    def export_to_excel(
        fundos: List[FundoData],
        incluir_metricas_avancadas: bool = True,
        incluir_alertas: bool = True
    ) -> bytes:
        """
        Exporta fundos para Excel com formatação

        Args:
            fundos: Lista de fundos
            incluir_metricas_avancadas: Incluir KPIs avançados
            incluir_alertas: Incluir aba de alertas

        Returns:
            Bytes do arquivo Excel
        """
        wb = Workbook()

        # Remover sheet padrão
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Aba 1: Resumo
        ExportService._criar_aba_resumo(wb, fundos)

        # Aba 2: Dados Detalhados
        ExportService._criar_aba_dados(wb, fundos, incluir_metricas_avancadas)

        # Aba 3: Alertas (se solicitado)
        if incluir_alertas:
            ExportService._criar_aba_alertas(wb, fundos)

        # Aba 4: Métricas Avançadas (se solicitado)
        if incluir_metricas_avancadas:
            ExportService._criar_aba_metricas_avancadas(wb, fundos)

        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def _criar_aba_resumo(wb: Workbook, fundos: List[FundoData]):
        """Cria aba de resumo executivo"""
        ws = wb.create_sheet("Resumo")

        # Título
        ws['A1'] = "Dashboard de Fundos - Resumo Executivo"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="005D90", end_color="005D90", fill_type="solid")
        ws.merge_cells('A1:D1')

        # Data de geração
        ws['A2'] = f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, color="666666")

        # Métricas agregadas
        ws['A4'] = "Métricas Gerais"
        ws['A4'].font = Font(size=14, bold=True)

        total_pl = sum(f.pl for f in fundos)
        total_caixa = sum(f.caixa_total for f in fundos)
        total_fundos = len(fundos)
        perc_caixa_pl = (total_caixa / total_pl * 100) if total_pl > 0 else 0

        metricas = [
            ("Total de Fundos", total_fundos, ""),
            ("Patrimônio Total", total_pl, "R$ {:,.2f}"),
            ("Caixa Total", total_caixa, "R$ {:,.2f}"),
            ("% Caixa/PL", perc_caixa_pl, "{:.2f}%"),
        ]

        row = 5
        for label, valor, fmt in metricas:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = fmt.format(valor) if fmt else valor
            row += 1

        # Distribuição por tipo
        ws[f'A{row + 1}'] = "Distribuição por Tipo"
        ws[f'A{row + 1}'].font = Font(size=14, bold=True)

        row += 2
        ws[f'A{row}'] = "Tipo"
        ws[f'B{row}'] = "Quantidade"
        ws[f'C{row}'] = "PL Total"
        ws[f'D{row}'] = "% do Total"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Agrupar por tipo
        fundos_por_tipo = {}
        for fundo in fundos:
            tipo = fundo.tipo or "Outros"
            if tipo not in fundos_por_tipo:
                fundos_por_tipo[tipo] = []
            fundos_por_tipo[tipo].append(fundo)

        row += 1
        for tipo in sorted(fundos_por_tipo.keys()):
            fundos_tipo = fundos_por_tipo[tipo]
            pl_tipo = sum(f.pl for f in fundos_tipo)
            ws[f'A{row}'] = tipo
            ws[f'B{row}'] = len(fundos_tipo)
            ws[f'C{row}'] = pl_tipo
            ws[f'C{row}'].number_format = 'R$ #,##0.00'
            ws[f'D{row}'] = (pl_tipo / total_pl * 100) if total_pl > 0 else 0
            ws[f'D{row}'].number_format = '0.00"%"'
            row += 1

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

    @staticmethod
    def _criar_aba_dados(wb: Workbook, fundos: List[FundoData], incluir_metricas: bool):
        """Cria aba com dados detalhados"""
        ws = wb.create_sheet("Dados Detalhados")

        # Header
        headers = [
            "Nome", "Tipo", "PL Atual", "PL D-1", "PL D-7", "PL D-30",
            "Caixa Total", "% Caixa/PL", "Var D-1", "Var D-7", "Var D-30"
        ]

        if incluir_metricas:
            headers.extend(["Sharpe Ratio", "Volatilidade", "Max Drawdown", "Health Score"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="005D90", end_color="005D90", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for row, fundo in enumerate(fundos, 2):
            ws.cell(row, 1, fundo.nome)
            ws.cell(row, 2, fundo.tipo)
            ws.cell(row, 3, fundo.pl).number_format = 'R$ #,##0.00'
            ws.cell(row, 4, fundo.pl_d1).number_format = 'R$ #,##0.00'
            ws.cell(row, 5, fundo.pl_d7).number_format = 'R$ #,##0.00'
            ws.cell(row, 6, fundo.pl_d30).number_format = 'R$ #,##0.00'
            ws.cell(row, 7, fundo.caixa_total).number_format = 'R$ #,##0.00'
            ws.cell(row, 8, fundo.perc_caixa_pl).number_format = '0.00"%"'
            ws.cell(row, 9, fundo.variacao_d1).number_format = '0.00"%"'
            ws.cell(row, 10, fundo.variacao_d7).number_format = '0.00"%"'
            ws.cell(row, 11, fundo.variacao_d30).number_format = '0.00"%"'

            # Colorir variações
            for col in [9, 10, 11]:
                cell = ws.cell(row, col)
                if cell.value and cell.value > 0:
                    cell.font = Font(color="10B981")
                elif cell.value and cell.value < 0:
                    cell.font = Font(color="EF4444")

            if incluir_metricas:
                metricas = calcular_metricas_fundo(fundo)
                ws.cell(row, 12, metricas['sharpe_ratio'])
                ws.cell(row, 13, metricas['volatilidade']).number_format = '0.00"%"'
                ws.cell(row, 14, metricas['max_drawdown']).number_format = '0.00"%"'
                ws.cell(row, 15, calcular_health_score(fundo))

        # Auto-ajustar colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Nome mais largo
        ws.column_dimensions['A'].width = 40

    @staticmethod
    def _criar_aba_alertas(wb: Workbook, fundos: List[FundoData]):
        """Cria aba de alertas"""
        ws = wb.create_sheet("Alertas")

        # Header
        headers = ["Fundo", "Tipo", "Nível", "Mensagem"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")

        row = 2
        for fundo in fundos:
            if fundo.tem_alertas():
                alertas = fundo.get_alertas()
                for nivel, mensagem in alertas:
                    ws.cell(row, 1, fundo.nome)
                    ws.cell(row, 2, fundo.tipo)
                    ws.cell(row, 3, nivel.upper())
                    ws.cell(row, 4, mensagem)

                    # Colorir por nível
                    color_map = {
                        "error": "FEE2E2",
                        "warning": "FEF3C7",
                        "info": "DBEAFE"
                    }
                    fill_color = color_map.get(nivel, "FFFFFF")
                    for col in range(1, 5):
                        ws.cell(row, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                    row += 1

        # Ajustar larguras
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 60

    @staticmethod
    def _criar_aba_metricas_avancadas(wb: Workbook, fundos: List[FundoData]):
        """Cria aba de métricas avançadas"""
        ws = wb.create_sheet("Métricas Avançadas")

        # Header
        headers = [
            "Nome", "Sharpe Ratio", "Sortino Ratio", "Volatilidade",
            "Max Drawdown", "VaR 95%", "CVaR 95%", "Calmar Ratio", "Health Score"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for row, fundo in enumerate(fundos, 2):
            metricas = calcular_metricas_fundo(fundo)

            ws.cell(row, 1, fundo.nome)
            ws.cell(row, 2, metricas['sharpe_ratio'])
            ws.cell(row, 3, metricas['sortino_ratio'])
            ws.cell(row, 4, metricas['volatilidade']).number_format = '0.00"%"'
            ws.cell(row, 5, metricas['max_drawdown']).number_format = '0.00"%"'
            ws.cell(row, 6, metricas['var_95']).number_format = '0.00"%"'
            ws.cell(row, 7, metricas['cvar_95']).number_format = '0.00"%"'
            ws.cell(row, 8, metricas['calmar_ratio'])
            ws.cell(row, 9, calcular_health_score(fundo))

            # Colorir health score
            health = calcular_health_score(fundo)
            cell = ws.cell(row, 9)
            if health >= 80:
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif health >= 60:
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif health < 40:
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        # Ajustar larguras
        ws.column_dimensions['A'].width = 40
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    @staticmethod
    def export_to_csv(fundos: List[FundoData]) -> bytes:
        """
        Exporta fundos para CSV

        Args:
            fundos: Lista de fundos

        Returns:
            Bytes do arquivo CSV
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "Nome", "Tipo", "PL Atual", "PL D-1", "PL D-7", "PL D-30",
            "Caixa Total", "% Caixa/PL", "Var D-1", "Var D-7", "Var D-30"
        ])

        # Dados
        for fundo in fundos:
            writer.writerow([
                fundo.nome,
                fundo.tipo,
                fundo.pl,
                fundo.pl_d1,
                fundo.pl_d7,
                fundo.pl_d30,
                fundo.caixa_total,
                fundo.perc_caixa_pl,
                fundo.variacao_d1,
                fundo.variacao_d7,
                fundo.variacao_d30
            ])

        # Converter para bytes
        output.seek(0)
        return output.getvalue().encode('utf-8-sig')  # BOM para Excel

    @staticmethod
    def get_filename(formato: str, prefixo: str = "dashboard") -> str:
        """
        Gera nome de arquivo com timestamp

        Args:
            formato: Extensão do arquivo (xlsx, csv, pdf)
            prefixo: Prefixo do nome

        Returns:
            Nome do arquivo com timestamp
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{prefixo}_{timestamp}.{formato}"


# Instância global
_export_service = ExportService()


def get_export_service() -> ExportService:
    """Retorna instância global do ExportService"""
    return _export_service
